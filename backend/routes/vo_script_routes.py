# backend/routes/vo_script_routes.py

from flask import Blueprint, request, jsonify, send_file, current_app
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm.attributes import flag_modified # Import flag_modified
import logging
import json # Added import
import os
from datetime import datetime, timezone # Import datetime utils
import io # For in-memory file handling
from openpyxl import Workbook # For creating Excel file
from openpyxl.styles import Font, Alignment, PatternFill # For formatting
from openpyxl.utils import get_column_letter # For setting column width
import re # Import regex for natural sort
import sqlalchemy as sa # Added import
from pydantic import BaseModel, Field, ValidationError
from typing import List, Optional, Dict, Any # Ensure Any is imported

# Assuming models and helpers are accessible, adjust imports as necessary
from backend import models # Added tasks import
from backend.models import get_db
from backend.utils.response_utils import make_api_response, model_to_dict # NEW imports
from backend import utils_openai # Import for direct OpenAI calls
from backend import utils_voscript # Import for DB utils
from backend.utils_prompts import _get_elevenlabs_rules # NEW IMPORT
from backend.tasks.script_tasks import run_script_collaborator_chat_task # Import the Celery task

vo_script_bp = Blueprint('vo_script_bp', __name__, url_prefix='/api')

# --- Helper function for natural sorting ---
def natural_sort_key(s):
    """Return a key for natural sorting (handles text and numbers)."""
    if not isinstance(s, str):
        return [s] # Handle non-strings gracefully
    # Split string into text and number parts
    return [int(text) if text.isdigit() else text.lower() 
            for text in re.split('([0-9]+)', s)]

@vo_script_bp.route('/vo-scripts', methods=['POST'])
def create_vo_script():
    """Creates a new VO script instance."""
    if not request.is_json:
        return make_api_response(error="Request must be JSON", status_code=400)
    data = request.get_json()
    name = data.get('name')
    template_id = data.get('template_id')
    character_description = data.get('character_description') # Now expecting a string

    # Validate required fields
    if not name or not template_id or character_description is None: # Allow empty description string
        return make_api_response(error="Missing required fields: name, template_id, character_description", status_code=400)
    
    # Validate template_id type
    try:
        template_id = int(template_id)
    except (ValueError, TypeError):
        return make_api_response(error="template_id must be an integer", status_code=400)
    
    # Validate character_description type (should be string)
    if not isinstance(character_description, str):
         return make_api_response(error="character_description must be a string", status_code=400)

    db: Session = None
    try:
        db = next(get_db())
        # Verify template exists
        template = db.query(models.VoScriptTemplate).get(template_id)
        if not template:
            return make_api_response(error=f"Template with ID {template_id} not found", status_code=404)
            
        new_vo_script = models.VoScript(
            name=name,
            template_id=template_id,
            character_description=character_description, # Pass string directly
            status='drafting' # Initial status
        )
        db.add(new_vo_script)
        
        # IMPORTANT: Also create placeholder vo_script_line entries for this new script
        # based on the lines defined in the associated template.
        template_lines = db.query(models.VoScriptTemplateLine).filter(
            models.VoScriptTemplateLine.template_id == template_id
        ).order_by(models.VoScriptTemplateLine.order_index).all() # Ensure order
        
        if not template_lines:
             logging.warning(f"Template ID {template_id} has no lines defined. Creating VO Script anyway.")
        
        vo_script_lines_to_add = []
        for t_line in template_lines:
            # Check if the template line has static_text content
            has_static_text = t_line.static_text is not None and t_line.static_text.strip() != ""
            
            # Create a new line with the template line's key copied over
            new_line = models.VoScriptLine(
                vo_script=new_vo_script, 
                template_line_id=t_line.id,
                # If static_text exists, copy it, mark as 'generated', and LOCK it
                # Otherwise, leave as 'pending' and unlocked for LLM generation
                status='generated' if has_static_text else 'pending',
                generated_text=t_line.static_text if has_static_text else None,
                line_key=t_line.line_key,  # Copy the line_key from the template line
                is_locked=has_static_text # NEW: Lock the line if it has static text
            )
            vo_script_lines_to_add.append(new_line)
            
        if vo_script_lines_to_add:
            db.add_all(vo_script_lines_to_add)
            
        db.commit()
        db.refresh(new_vo_script)
        logging.info(f"Created VO script ID {new_vo_script.id} ('{name}') using template ID {template_id}, added {len(vo_script_lines_to_add)} pending lines.")
        # Include lines in the response? Maybe not for POST, keep it lean.
        # Fetch again to include template name?
        created_script = db.query(models.VoScript).options(joinedload(models.VoScript.template)).get(new_vo_script.id)
        resp_data = model_to_dict(created_script)
        resp_data['template_name'] = created_script.template.name if created_script.template else None
        return make_api_response(data=resp_data, status_code=201)
        
    except IntegrityError as e:
        db.rollback()
        # Could be FK violation if template disappears, or duplicate name if unique constraint added
        logging.exception(f"Database integrity error creating vo_script: {e}")
        # Check for unique constraint violation specifically? Depends on DB schema
        # if "UNIQUE constraint failed" in str(e):
        #     return make_api_response(error=f"VO Script with name '{name}' already exists.", status_code=409)
        return make_api_response(error="Database error creating script.", status_code=500)
    except Exception as e:
        db.rollback()
        logging.exception(f"Error creating VO script: {e}")
        return make_api_response(error="Failed to create VO script", status_code=500)
    finally:
        if db and db.is_active: db.close()

@vo_script_bp.route('/vo-scripts', methods=['GET'])
def list_vo_scripts():
    """Lists all VO script instances."""
    db: Session = None
    try:
        db = next(get_db())
        # Eager load template name for display
        scripts = db.query(models.VoScript).options(
            joinedload(models.VoScript.template)
        ).order_by(models.VoScript.updated_at.desc()).all()
        
        script_list = []
        for script in scripts:
            s_dict = model_to_dict(script, ['id', 'name', 'template_id', 'status', 'updated_at', 'character_description', 'created_at']) # Added missing fields
            # Add template name if loaded
            s_dict['template_name'] = script.template.name if script.template else None
            script_list.append(s_dict)
            
        logging.info(f"Returning {len(script_list)} VO scripts.")
        return make_api_response(data=script_list)
    except Exception as e:
        logging.exception(f"Error listing VO scripts: {e}")
        return make_api_response(error="Failed to list VO scripts", status_code=500)
    finally:
        if db and db.is_active: db.close()

@vo_script_bp.route('/vo-scripts/<int:script_id>', methods=['GET'])
def get_vo_script(script_id):
    """Gets details for a specific VO script instance, including its lines and refinement prompts."""
    db: Session = None
    try:
        db = next(get_db())
        # Eager load related data: template info, lines, template lines, and categories
        script = db.query(models.VoScript).options(
            joinedload(models.VoScript.template).selectinload(models.VoScriptTemplate.categories), # Load template and its categories
            selectinload(models.VoScript.lines).selectinload(models.VoScriptLine.template_line) # Load lines and their template line link
        ).get(script_id)
        
        if script is None:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)
            
        # Serialize the script, including the new refinement_prompt
        script_data = model_to_dict(script) 
        if script.template:
            script_data['template_name'] = script.template.name
            script_data['template_description'] = script.template.description 
            script_data['template_prompt_hint'] = script.template.prompt_hint 
            # Explicitly add categories with their refinement prompts
            template_categories = {
                c.id: {
                    "id": c.id,
                    "name": c.name,
                    "instructions": c.prompt_instructions,
                    "refinement_prompt": c.refinement_prompt
                }
                for c in script.template.categories
            }
        else:
             template_categories = {}
            
        lines_by_category = {}
        # REMOVED: Initial sorting by order_index. We process all lines and group first.
        # lines_with_sort_key = [] 
        # ... (code to prepare lines_with_sort_key) ...
        # ordered_lines_data = sorted(lines_with_sort_key, key=lambda item: item["sort_order"])
        
        # --- Process ALL lines --- #
        # MODIFIED: Iterate directly through script.lines
        for line in script.lines: 
            # Corrected: Pass include list as positional argument
            line_dict_base = model_to_dict(line, [
                'id', 'generated_text', 'status', 'latest_feedback', 
                'generation_history', 'is_locked', 'template_line_id',
                'created_at', 'updated_at'
            ])

            # Explicitly and safely get other attributes
            db_category_id = getattr(line, 'category_id', None)
            db_line_key = getattr(line, 'line_key', None)
            db_order_index = getattr(line, 'order_index', None)
            db_prompt_hint = getattr(line, 'prompt_hint', None)
            
            template_line_key = line.template_line.line_key if line.template_line else None
            template_order_index = line.template_line.order_index if line.template_line and line.template_line.order_index is not None else None
            template_line_category_id = line.template_line.category_id if line.template_line else None
            template_prompt_hint = line.template_line.prompt_hint if line.template_line else None

            # Final determination logic (Prioritize direct values)
            # Use a fallback key if none is found
            final_line_key = db_line_key or template_line_key or f'line_{line.id}'
            final_order_index = db_order_index if db_order_index is not None else template_order_index
            final_prompt_hint = db_prompt_hint # Direct hint is primary
            final_template_prompt_hint = template_prompt_hint # Hint from template
            final_category_id = db_category_id or template_line_category_id # Prioritize direct ID
                
            # Combine into the final dictionary
            line_dict = {
                **line_dict_base,
                'line_key': final_line_key, 
                'order_index': final_order_index,
                'prompt_hint': final_prompt_hint, 
                'template_prompt_hint': final_template_prompt_hint, 
                'category_id': final_category_id
            }
            
            # Ensure datetimes are strings 
            line_dict['created_at'] = line_dict['created_at'].isoformat() if line_dict['created_at'] and hasattr(line_dict['created_at'], 'isoformat') else line_dict['created_at']
            line_dict['updated_at'] = line_dict['updated_at'].isoformat() if line_dict['updated_at'] and hasattr(line_dict['updated_at'], 'isoformat') else line_dict['updated_at']

            # --- Determine grouping category --- 
            category_id_for_grouping = line_dict.get('category_id')
            logging.debug(f"Line ID {line.id} (Key: {line_dict.get('line_key')}): Category ID for grouping = {category_id_for_grouping}") # LOG 1
            
            # --- SIMPLIFIED CATEGORY LOOKUP ---
            category_name = "Uncategorized" # Default
            category_data_id = None
            category_instructions = None
            category_refinement_prompt = None

            if category_id_for_grouping:
                # Attempt to find category info directly using the ID
                # Check the preloaded dict first for efficiency (for template lines)
                category_info = template_categories.get(category_id_for_grouping)
                if category_info:
                    category_name = category_info['name']
                    category_data_id = category_info['id']
                    category_instructions = category_info['instructions']
                    category_refinement_prompt = category_info['refinement_prompt']
                    logging.debug(f"Line ID {line.id}: Found category '{category_name}' (ID: {category_data_id}) in pre-loaded template_categories.") # LOG 2a
                else:
                    # If not in preloaded dict (likely a custom line), query DB directly
                    logging.debug(f"Line ID {line.id}: category_id {category_id_for_grouping} not in template_categories. Querying DB directly...") # LOG 2b
                    category_obj = db.query(
                        models.VoScriptTemplateCategory.id, 
                        models.VoScriptTemplateCategory.name, 
                        models.VoScriptTemplateCategory.prompt_instructions, 
                        models.VoScriptTemplateCategory.refinement_prompt
                    ).filter(models.VoScriptTemplateCategory.id == category_id_for_grouping).first()
                    
                    if category_obj:
                        category_name = category_obj.name
                        category_data_id = category_obj.id
                        category_instructions = category_obj.prompt_instructions
                        category_refinement_prompt = category_obj.refinement_prompt
                        logging.debug(f"Line ID {line.id}: Found category '{category_obj.name}' (ID: {category_data_id}) via direct DB query.") # LOG 2c
                    else:
                        logging.warning(f"Line ID {line.id}: Could not find category details via direct DB query for existing category ID: {category_id_for_grouping}. Falling back to Uncategorized.") # LOG 2d
                        # category_name remains "Uncategorized"
            else:
                 logging.debug(f"Line ID {line.id}: No category ID found.") # LOG 3 (Renumbered)

            # Grouping logic - Initialize category if new
            if category_name not in lines_by_category:
                 lines_by_category[category_name] = {
                     "id": category_data_id, 
                     'name': category_name,
                     'instructions': category_instructions, 
                     'refinement_prompt': category_refinement_prompt, 
                     'lines': []
                 }
                 # Add category details if newly created (might be redundant if always querying?)
                 if category_data_id and category_name != "Uncategorized":
                     logging.debug(f"Initialized category group '{category_name}' (ID: {category_data_id})")
                 else:
                      logging.debug(f"Initialized category group '{category_name}'")
            
            # Update existing group if needed
            if category_data_id and lines_by_category[category_name].get('id') is None:
                lines_by_category[category_name]['id'] = category_data_id
                lines_by_category[category_name]['instructions'] = category_instructions
                lines_by_category[category_name]['refinement_prompt'] = category_refinement_prompt
                logging.debug(f"Updated existing group '{category_name}' with details (ID: {category_data_id}) found via direct query.")

            lines_by_category[category_name]['lines'].append(line_dict)
            logging.debug(f"Line ID {line.id}: Appended to category '{category_name}'.") # LOG 4 (Renumbered)

        # --- NEW: Post-processing Sort --- #
        # 1. Sort lines within each category by natural sort of line_key
        for category_name in lines_by_category:
            lines_by_category[category_name]['lines'].sort(key=lambda line: natural_sort_key(line.get('line_key')))
        
        # 2. Sort the categories themselves alphabetically by name
        sorted_categories = sorted(lines_by_category.values(), key=lambda cat: cat['name'])

        # Assign the fully sorted structure to the response
        script_data['categories'] = sorted_categories
            
        return make_api_response(data=script_data)
    except Exception as e:
        logging.exception(f"Error getting VO script {script_id}: {e}")
        return make_api_response(error="Failed to get VO script", status_code=500)
    finally:
        if db and db.is_active: db.close()

@vo_script_bp.route('/vo-scripts/<int:script_id>', methods=['PUT'])
def update_vo_script(script_id):
    """Updates an existing VO script instance (name, char desc, status, refinement_prompt)."""
    if not request.is_json:
        return make_api_response(error="Request must be JSON", status_code=400)
    data = request.get_json()
    
    db: Session = None
    try:
        db = next(get_db())
        script = db.query(models.VoScript).get(script_id)
        if script is None:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)

        updated = False
        # --- Update Name --- #
        if 'name' in data:
            new_name = data['name']
            if not new_name:
                 return make_api_response(error="Name cannot be empty", status_code=400)
            if new_name != script.name:
                 script.name = new_name
                 updated = True
        
        # --- Update Character Description --- #         
        if 'character_description' in data:
            new_desc = data['character_description']
            if not isinstance(new_desc, str):
                 return make_api_response(error="character_description must be a string", status_code=400)
            if new_desc != script.character_description:
                 script.character_description = new_desc
                 updated = True
        
        # --- Update Refinement Prompt --- #
        if 'refinement_prompt' in data:
             new_prompt = data['refinement_prompt']
             if not isinstance(new_prompt, (str, type(None))):
                 return make_api_response(error="refinement_prompt must be a string or null", status_code=400)
             if new_prompt != script.refinement_prompt:
                  script.refinement_prompt = new_prompt
                  updated = True
        
        # --- Update Status --- #
        if 'status' in data:
            new_status = data['status']
            allowed_statuses = ['drafting', 'review', 'locked'] # Example allowed statuses
            if new_status not in allowed_statuses:
                return make_api_response(error=f"Invalid status '{new_status}'. Allowed: {allowed_statuses}", status_code=400)
            if new_status != script.status:
                 script.status = new_status
                 updated = True

        # If nothing was actually changed, return the current script data 
        if not updated:
             # Return the basic script data (including refinement_prompt)
             return make_api_response(data=model_to_dict(script))

        # Commit changes if any were made
        db.commit()
        db.refresh(script) # Refresh to get updated timestamp etc.
        logging.info(f"Updated VO script ID {script.id}")
        
        # Return the updated basic script data (client can refetch full details if needed)
        return make_api_response(data=model_to_dict(script))

    except Exception as e:
        db.rollback()
        logging.exception(f"Error updating VO script {script_id}: {e}")
        return make_api_response(error="Failed to update VO script", status_code=500)
    finally:
        if db and db.is_active: db.close()

@vo_script_bp.route('/vo-scripts/<int:script_id>', methods=['DELETE'])
def delete_vo_script(script_id):
    """Deletes a VO script instance and its associated lines."""
    db: Session = None
    try:
        db = next(get_db())
        script = db.query(models.VoScript).get(script_id)
        if script is None:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)
        
        script_name = script.name # Get name for logging
        db.delete(script) # Cascade should delete associated VoScriptLine records
        db.commit()
        logging.info(f"Deleted VO script ID {script_id} (Name: '{script_name}')")
        return make_api_response(data={"message": f"VO Script '{script_name}' deleted successfully"})
    except Exception as e:
        db.rollback()
        logging.exception(f"Error deleting VO script {script_id}: {e}")
        return make_api_response(error="Failed to delete VO script", status_code=500)
    finally:
        if db and db.is_active: db.close()

# --- VoScript Agent Trigger --- #

@vo_script_bp.route('/vo-scripts/<int:script_id>/run-agent', methods=['POST'])
def run_vo_script_agent(script_id):
    """Triggers the VO Script creation/refinement agent via Celery."""
    if not request.is_json:
        return make_api_response(error="Request must be JSON", status_code=400)
    data = request.get_json()
    task_type = data.get('task_type', 'generate_draft') 
    feedback_data = data.get('feedback') # Optional feedback dict/list for refine_feedback
    category_name = data.get('category_name') # Optional category name for refine_category
    target_model = data.get('model') # Optional model override
    
    # Define allowed task types
    allowed_task_types = ['generate_draft', 'refine_feedback', 'refine_category']
    
    if task_type not in allowed_task_types:
        return make_api_response(error=f"Invalid task_type. Allowed: {allowed_task_types}", status_code=400)
        
    # Validate category_name if task_type requires it
    if task_type == 'refine_category' and not category_name:
         return make_api_response(error="Missing required field 'category_name' for task_type 'refine_category'", status_code=400)
    
    # Ensure category_name is string if provided
    if category_name and not isinstance(category_name, str):
         return make_api_response(error="'category_name' must be a string if provided", status_code=400)
        
    db: Session = None
    db_job = None
    try:
        db = next(get_db())
        # Verify script exists
        script = db.query(models.VoScript).get(script_id)
        if not script:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)
            
        # For 'generate_draft' task type, use our improved category-based generation
        if task_type == 'generate_draft':
            # Get distinct categories from this script's template
            if script.template_id:
                template = db.query(models.VoScriptTemplate).get(script.template_id)
                if template:
                    categories = db.query(models.VoScriptTemplateCategory).filter(
                        models.VoScriptTemplateCategory.template_id == template.id,
                        models.VoScriptTemplateCategory.is_deleted == False
                    ).all()
                    
                    if categories:
                        # Process all categories using our improved batch approach
                        results = {'categories': []}
                        
                        # Create main job record
                        job_params = {
                            "task_type": "generate_draft_by_categories",
                            "categories_count": len(categories)
                        }
                        if target_model:
                            job_params["model"] = target_model
                            
                        db_job = models.GenerationJob(
                            status="PENDING",
                            job_type="script_creation", 
                            target_batch_id=str(script_id),
                            parameters_json=json.dumps(job_params)
                        )
                        db.add(db_job)
                        db.commit()
                        db.refresh(db_job)
                        db_job_id = db_job.id
                        
                        # For each category, trigger a category batch generation task
                        for category in categories:
                            try:
                                # Create child job for this category
                                category_job_params = {
                                    "task_type": "generate_category",
                                    "category_name": category.name,
                                    "parent_job_id": db_job_id
                                }
                                if target_model:
                                    category_job_params["model"] = target_model
                                
                                category_job = models.GenerationJob(
                                    status="PENDING",
                                    job_type="script_creation", 
                                    target_batch_id=str(script_id),
                                    parameters_json=json.dumps(category_job_params)
                                )
                                db.add(category_job)
                                db.flush()  # Get ID without committing transaction
                                category_job_id = category_job.id
                                
                                # Enqueue task for this category
                                task = tasks.generate_category_lines.delay(
                                    category_job_id,
                                    script_id,
                                    category.name,
                                    target_model
                                )
                                
                                # Update job with task ID
                                category_job.celery_task_id = task.id
                                
                                # Track in results
                                results['categories'].append({
                                    'category_name': category.name,
                                    'job_id': category_job_id,
                                    'task_id': task.id
                                })
                                
                                logging.info(f"Enqueued category generation task for '{category.name}': Job ID {category_job_id}, Task ID {task.id}")
                            except Exception as e:
                                logging.exception(f"Error enqueuing generation for category '{category.name}': {e}")
                        
                        # Commit all child jobs
                        db.commit()
                        
                        # Update main job status
                        db_job.result_message = f"Triggered generation for {len(categories)} categories."
                        db.commit()
                        
                        return make_api_response(
                            data={
                                'job_id': db_job_id, 
                                'task_type': 'generate_draft_by_categories',
                                'results': results
                            }, 
                            status_code=202
                        )
            
            # If we got here, either:
            # 1. The script has no template
            # 2. The template has no categories
            # 3. There was some error getting categories
            # In these cases, fall back to the legacy agent-based approach
            logging.warning(f"Falling back to legacy agent for script {script_id} (No categories found)")
        
        # For all other task types or fallback, use the original agent-based approach
        # Create Job record
        job_params = {"task_type": task_type} # Base params
        if feedback_data:
             job_params["feedback"] = feedback_data
        if category_name:
             job_params["category_name"] = category_name
        if target_model:
             job_params["model"] = target_model
             
        db_job = models.GenerationJob(
            status="PENDING",
            job_type="script_creation", 
            target_batch_id=str(script_id), 
            parameters_json=json.dumps(job_params) # Store task type, feedback, category
        )
        db.add(db_job)
        db.commit()
        db.refresh(db_job)
        db_job_id = db_job.id
        logging.info(f"Created Script Creation Job DB ID: {db_job_id} for VO Script ID {script_id}")

        # Enqueue Celery task
        task = tasks.run_script_creation_agent.delay(
            db_job_id,
            script_id,
            task_type,
            feedback_data, # Pass feedback (can be None)
            category_name # Pass category name (can be None)
        )
        logging.info(f"Enqueued script creation task: Celery ID {task.id}, DB Job ID {db_job_id}")
        
        # Link Celery task ID to Job record
        db_job.celery_task_id = task.id
        db.commit()
        
        return make_api_response(data={'job_id': db_job_id, 'task_id': task.id}, status_code=202)
        
    except Exception as e:
        logging.exception(f"Error submitting script creation job for VO script {script_id}: {e}")
        if db_job and db_job.id and db.is_active: # Mark DB job as failed if possible
            try: 
                db_job.status = "SUBMIT_FAILED"
                db_job.result_message = f"Enqueue failed: {e}"
                db.commit()
            except: 
                db.rollback()
        elif db and db.is_active:
             db.rollback()
        return make_api_response(error="Failed to start script creation task", status_code=500)
    finally:
        if db and db.is_active: db.close()

# --- VoScript Feedback --- #

@vo_script_bp.route('/vo-scripts/<int:script_id>/feedback', methods=['POST'])
def submit_vo_script_feedback(script_id):
    """Submits user feedback for a specific line within a VO script."""
    if not request.is_json:
        return make_api_response(error="Request must be JSON", status_code=400)
    data = request.get_json()
    line_id = data.get('line_id')
    feedback_text = data.get('feedback_text')

    if line_id is None or feedback_text is None: # Allow empty feedback text
        return make_api_response(error="Missing required fields: line_id, feedback_text", status_code=400)
    
    try:
        line_id = int(line_id)
    except (ValueError, TypeError):
        return make_api_response(error="line_id must be an integer", status_code=400)
        
    if not isinstance(feedback_text, str):
        return make_api_response(error="feedback_text must be a string", status_code=400)

    db: Session = None
    try:
        db = next(get_db())
        
        # Find the specific VoScriptLine using both script_id and line_id
        script_line = db.query(models.VoScriptLine).filter(
            models.VoScriptLine.vo_script_id == script_id,
            models.VoScriptLine.id == line_id
        ).first()

        if not script_line:
            # Check if the script itself exists to provide a better error message
            script_exists = db.query(models.VoScript.id).filter(models.VoScript.id == script_id).scalar() is not None
            if not script_exists:
                return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)
            else:
                return make_api_response(error=f"Line with ID {line_id} not found within VO Script ID {script_id}", status_code=404)

        # Update the feedback field
        script_line.latest_feedback = feedback_text
        
        # Potentially update script status? Or just update the line?
        # For now, just update the line.
        
        db.commit()
        db.refresh(script_line) # Refresh to get updated timestamp if any
        logging.info(f"Updated feedback for VO Script ID {script_id}, Line ID {line_id}")
        
        # Return the updated line info (or just success)
        return make_api_response(data=model_to_dict(script_line, ['id', 'vo_script_id', 'latest_feedback', 'updated_at']))

    except Exception as e:
        db.rollback()
        logging.exception(f"Error submitting feedback for VO script {script_id}, line {line_id}: {e}")
        return make_api_response(error="Failed to submit feedback", status_code=500)
    finally:
        if db and db.is_active: db.close()

# --- NEW: Line Refinement Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/lines/<int:line_id>/refine', methods=['POST'])
def refine_vo_script_line(script_id: int, line_id: int):
    from backend.app import model_to_dict 
    """Refines a single VO script line using OpenAI based on user prompt,
       optionally applying ElevenLabs best practices.
    """
    data = request.get_json()
    if not data:
        return jsonify({"error": "Missing request body"}), 400
    
    user_prompt = data.get('line_prompt')
    apply_best_practices = data.get('apply_best_practices', False) # NEW: Get checkbox value
    target_model = data.get('model', utils_openai.DEFAULT_REFINEMENT_MODEL)

    if not user_prompt and not apply_best_practices:
        logging.warning(f"Refine line request missing 'line_prompt' and apply_best_practices is false for script {script_id}, line {line_id}")
        return jsonify({"error": "Missing 'line_prompt' or 'apply_best_practices' must be true"}), 400

    db: Session = next(get_db())
    try:
        # --- ADD LOCK CHECK --- #
        line = db.query(models.VoScriptLine).filter(
            models.VoScriptLine.id == line_id,
            models.VoScriptLine.vo_script_id == script_id
        ).first()
        
        if not line:
            # Use standard response format
            return make_api_response(error=f"Line context not found for line_id {line_id}", status_code=404)
        
        if line.is_locked:
            logging.info(f"Skipping refinement for locked line {line_id} in script {script_id}.")
            # Return current line data instead of error
            return make_api_response(data=model_to_dict(line), status_code=200) 
        # --- END LOCK CHECK --- #

        # 1. Get context for the line (already fetched the line object)
        # We still need the full context for the prompt
        line_context = utils_voscript.get_line_context(db, line_id) 
        if not line_context: # Should ideally not happen if line exists, but good safeguard
            logging.warning(f"Refine line request: Context could not be built for script {script_id}, line {line_id}")
            return make_api_response(error=f"Line context could not be built for line_id {line_id}", status_code=500)

        # --- 2. Construct prompt for OpenAI --- 
        prompt_parts = []
        prompt_parts.append("You are a creative writer for video game voiceovers.")
        prompt_parts.append(f"Character Description:\n{line_context.get('character_description', 'N/A')}\n")
        prompt_parts.append(f"Template Hint: {line_context.get('template_hint', 'N/A')}")
        prompt_parts.append(f"Category: {line_context.get('category_name', 'N/A')}")
        prompt_parts.append(f"Category Instructions: {line_context.get('category_instructions', 'N/A')}")
        prompt_parts.append(f"Line Key: {line_context.get('line_key', 'N/A')}")
        prompt_parts.append(f"Line Hint: {line_context.get('line_template_hint', 'N/A')}\n")
        prompt_parts.append(f"Current Line Text:\n{line_context.get('current_text', '')}")
        
        elevenlabs_rules = None
        if apply_best_practices:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            prompt_file_dir = os.path.dirname(current_dir)
            rules_path = os.path.join(prompt_file_dir, 'prompts', 'scripthelp.md')
            elevenlabs_rules = _get_elevenlabs_rules(rules_path)
            if not elevenlabs_rules:
                logging.warning(f"Could not load ElevenLabs rules from {rules_path}. Proceeding without them.")

        if apply_best_practices and elevenlabs_rules:
            # Construct two-stage prompt using concatenation/joining
            user_request_text = user_prompt if user_prompt else "No specific user refinement request provided. Focus only on applying ElevenLabs best practices."
            prompt_parts.append("\n\n--- Stage 1: User Refinement Request ---")
            prompt_parts.append(f"User Request: \"{user_request_text}\"\n")
            prompt_parts.append("--- Stage 2: Apply ElevenLabs Best Practices ---")
            prompt_parts.append(f"ElevenLabs Rules:\n{elevenlabs_rules}\n")
            prompt_parts.append("--- Instructions ---")
            prompt_parts.append("1. Rewrite the 'Current Line Text' based *only* on the 'User Refinement Request' and the original context (Character Description, Hints, etc.). If no user request was provided, keep the 'Current Line Text' as is for this stage.")
            prompt_parts.append("2. Take the result from step 1 and apply the 'ElevenLabs Rules' to it, adding appropriate formatting like <break time=\"0.5s\"/> tags for pauses, and potentially phonetic hints or emphasis based on the rules and the refined text's meaning.")
            prompt_parts.append("3. Output ONLY the final text after applying both stages. Do not include explanations or intermediate steps.\n")
            prompt_parts.append("--- FINAL REFINED AND FORMATTED TEXT ---")
        else:
            # Construct original single-stage prompt
            user_request_text = user_prompt if user_prompt else "No refinement request provided. Check if any change is implied by context, otherwise return original text."
            prompt_parts.append(f"\nUser Refinement Request: \"{user_request_text}\"\n")
            prompt_parts.append("Rewrite the 'Current Line Text' based ONLY on the 'User Refinement Request', while staying consistent with the character description and other hints.")
            prompt_parts.append("Only output the refined line text, with no extra explanation or preamble. If no change is needed based on the request, output the original text exactly.")

        openai_prompt = "\n".join(prompt_parts)
        logging.debug(f"Sending prompt to OpenAI for line {line_id} (Apply Rules: {apply_best_practices}). Prompt start: {openai_prompt[:250]}...") # Increased log length

        # 3. Call OpenAI Responses API
        refined_text = utils_openai.call_openai_responses_api(
            prompt=openai_prompt,
            model=target_model,
        )
        
        if refined_text is None:
            logging.error(f"OpenAI refinement failed for script {script_id}, line {line_id}")
            return jsonify({"error": "OpenAI refinement failed. Check logs."}), 500
            
        logging.info(f"Refined text received for line {line_id}: '{refined_text[:100]}...'")

        # 4. Update Database
        # Determine new status - simple logic for now
        new_status = "review" if refined_text != line_context.get('current_text') else line_context.get('status', 'generated')
        
        updated_line = utils_voscript.update_line_in_db(
            db, 
            line_id, 
            refined_text, 
            new_status, 
            target_model # Log which model did the refinement
        )
        
        if updated_line is None:
            logging.error(f"Database update failed after refinement for script {script_id}, line {line_id}")
            # DB already rolled back in utility function
            return jsonify({"error": "Database update failed after refinement."}), 500

        # 5. Return updated line data
        # Use model_to_dict or similar serialization
        return jsonify({"data": model_to_dict(updated_line)}), 200

    except Exception as e:
        # Log any unexpected errors during the process
        logging.exception(f"Unexpected error during line refinement for script {script_id}, line {line_id}: {e}")
        # Ensure DB session is rolled back if error occurred before update_line_in_db handled it
        if db.is_active:
             try: db.rollback()
             except: pass # Ignore rollback errors
        return jsonify({"error": "An unexpected error occurred during refinement."}), 500
    finally:
        # Ensure DB session is closed
        if db:
            db.close()

# --- NEW: Category Refinement Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/categories/refine', methods=['POST'])
def refine_vo_script_category(script_id: int):
    """Refines all VO script lines within a category using OpenAI,
       optionally applying ElevenLabs best practices.
    """
    from backend.app import model_to_dict # Import locally
    data = request.get_json()
    if not data:
        return jsonify({"error": "Missing request body"}), 400
    
    category_name = data.get('category_name')
    category_prompt = data.get('category_prompt')
    apply_best_practices = data.get('apply_best_practices', False) # NEW: Get flag
    target_model = data.get('model', utils_openai.DEFAULT_REFINEMENT_MODEL)

    if not category_name:
        return jsonify({"error": "Missing 'category_name' in request body"}), 400
    # Allow empty category_prompt if apply_best_practices is true
    if not category_prompt and not apply_best_practices:
         return jsonify({"error": "Missing 'category_prompt' or 'apply_best_practices' must be true"}), 400

    db: Session = next(get_db())
    updated_lines_data = []
    errors_occurred = False
    
    # --- NEW: Get ElevenLabs rules if requested --- 
    elevenlabs_rules = None
    if apply_best_practices:
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            prompt_file_dir = os.path.dirname(current_dir)
            rules_path = os.path.join(prompt_file_dir, 'prompts', 'scripthelp.md')
            elevenlabs_rules = _get_elevenlabs_rules(rules_path)
            if not elevenlabs_rules:
                logging.warning(f"Could not load ElevenLabs rules from {rules_path} for category refine. Proceeding without them.")
        except Exception as e:
            logging.exception(f"Error loading ElevenLabs rules for category refine: {e}")
            # Decide: fail the request or proceed without rules? Proceeding for now.
            elevenlabs_rules = None # Ensure it's None if loading failed
    # --- END NEW --- 
    
    try:
        # 1. Get context (now includes script prompt)
        lines_to_process = utils_voscript.get_category_lines_context(db, script_id, category_name)
        
        if not lines_to_process:
            logging.info(f"No lines found for category '{category_name}' in script {script_id}. Nothing to refine.")
            return jsonify({"data": []}), 200 

        logging.info(f"Found {len(lines_to_process)} potential lines to refine for category '{category_name}' in script {script_id}.")

        # 2. Iterate and refine each NON-LOCKED line
        # --- NEW: Create map for quick text lookup --- 
        line_texts_map = {ltx['line_id']: ltx.get('current_text', '') for ltx in lines_to_process}
        
        for line_index, line_context in enumerate(lines_to_process):
            line_id = line_context['line_id']
            
            if line_context.get('is_locked', False):
                logging.info(f"Skipping locked line {line_id} during category refinement.")
                continue 
                
            # --- Construct Prompt (Conditional) --- 
            # Base context (Always included)
            base_context_prompt_parts = [
                f"You are a creative writer for video game voiceovers.",
                f"Character Description:\n{line_context.get('character_description', 'N/A')}\n",
                f"Template Hint: {line_context.get('template_hint', 'N/A')}",
                f"Category: {line_context.get('category_name', 'N/A')}",
                f"Category Instructions: {line_context.get('category_instructions', 'N/A')}",
                f"Line Key: {line_context.get('line_key', 'N/A')}",
                f"Line Hint: {line_context.get('line_template_hint', 'N/A')}\n",
                f"Current Line Text:\n{line_context.get('current_text', '')}"
            ]
            
            # User refinement prompts (Always included, adjusted if empty)
            script_prompt_text = line_context.get('script_refinement_prompt') or "N/A"
            line_feedback_text = line_context.get('latest_feedback') or "N/A"
            category_request_prompt = category_prompt or "N/A"
            user_request_summary = f"Global Script Prompt: {script_prompt_text}\nCategory Prompt: {category_request_prompt}\nLine Feedback/Prompt: {line_feedback_text}"
            user_request_text_for_stage1 = category_prompt if category_prompt else "No specific category refinement request provided."
            
            # Sibling examples (Calculate only if needed)
            sibling_examples_text_parts = []
            variety_instruction = "" # Initialize as empty
            if len(lines_to_process) > 1:
                sibling_examples = []
                example_count = 0
                for i, sibling_context in enumerate(lines_to_process):
                    if i == line_index or example_count >= 5:
                        continue
                    sibling_key = sibling_context.get('line_key') or f"line_{sibling_context['line_id']}"
                    sibling_text = line_texts_map.get(sibling_context['line_id'], "[ERROR: Text not found]")
                    if sibling_text:
                         sibling_examples.append(f"- {sibling_key}: \"{sibling_text}\"") 
                         example_count += 1
                if sibling_examples:
                    sibling_examples_text_parts.append("\n--- Sibling Line Examples (for context and ensuring variety) ---")
                    sibling_examples_text_parts.extend(sibling_examples)
                    # Define variety instruction ONLY if siblings exist
                    variety_instruction = "\n\nIMPORTANT: Ensure the refined output for this specific line is varied and distinct (e.g., in structure, theme, punchline) compared to the Sibling Line Examples provided. Avoid repetitive patterns."

            # ElevenLabs rules (Conditional)
            elevenlabs_rules_parts = []
            if apply_best_practices and elevenlabs_rules:
                elevenlabs_rules_parts.append("\n--- Stage 2: Apply ElevenLabs Best Practices ---")
                elevenlabs_rules_parts.append(f"ElevenLabs Rules:\n{elevenlabs_rules}")

            # Final Instructions (Conditional)
            final_instructions_parts = []
            if apply_best_practices and elevenlabs_rules:
                 final_instructions_parts.append("\n--- Instructions ---")
                 final_instructions_parts.append(f"1. Rewrite the 'Current Line Text' based *only* on the hierarchical user refinement prompts above ({user_request_summary}) and the original context (Character Description, Hints, etc.).")
                 final_instructions_parts.append(f"2. Take the result from step 1 and apply the 'ElevenLabs Rules' to it, adding appropriate formatting like <break time=\"0.5s\"/> tags for pauses, etc.{variety_instruction}")
                 final_instructions_parts.append("3. Output ONLY the final text after applying both stages. Do not include explanations or intermediate steps.\n")
                 final_instructions_parts.append("--- FINAL REFINED AND FORMATTED TEXT ---")
            else:
                 final_instructions_parts.append(f"\n\n--- Prompts (Apply these hierarchically) ---\n{user_request_summary}")
                 final_instructions_parts.append(f"\nRewrite the 'Current Line Text' based on ALL applicable prompts above (Global, Category, Line), while staying consistent with the character description and other hints. Prioritize the most specific prompt if conflicts arise.{variety_instruction}")
                 final_instructions_parts.append("Only output the refined line text, with no extra explanation or preamble. If no change is needed based on the prompts, output the original text exactly.")

            # Combine the parts
            openai_prompt_list = base_context_prompt_parts + sibling_examples_text_parts
            # Add Stage 1 marker only if doing two stages
            if apply_best_practices and elevenlabs_rules:
                openai_prompt_list.append(f"\n\n--- Stage 1: User Refinement Request ---\nUser Request: \"{user_request_text_for_stage1}\"")
            openai_prompt_list.extend(elevenlabs_rules_parts)
            openai_prompt_list.extend(final_instructions_parts)
            
            openai_prompt = "\n".join(openai_prompt_list)
            # --- END REVISED Prompt Construction --- 
            
            logging.debug(f"Sending category-refine prompt to OpenAI for line {line_id} (Apply Rules: {apply_best_practices})...")
            refined_text = utils_openai.call_openai_responses_api(
                prompt=openai_prompt,
                model=target_model
            )

            if refined_text is None:
                logging.error(f"OpenAI category refinement failed for script {script_id}, line {line_id}")
                errors_occurred = True 
                continue 
            
            logging.info(f"Refined text received for line {line_id} via category refine.")

            # Update Database for this line
            new_status = "review" if refined_text != line_context.get('current_text') else line_context.get('status', 'generated')
            updated_line = utils_voscript.update_line_in_db(
                db, line_id, refined_text, new_status, target_model
            )
            
            if updated_line is None:
                logging.error(f"Database update failed after category refinement for script {script_id}, line {line_id}")
                errors_occurred = True 
            else:
                updated_lines_data.append(model_to_dict(updated_line)) 

        # 3. Return results
        if errors_occurred:
             status_code = 207 # Multi-Status
             message = "Category refinement completed with some errors."
        else:
             status_code = 200
             message = "Category refinement completed successfully."
             
        return jsonify({"message": message, "data": updated_lines_data}), status_code

    except Exception as e:
        logging.exception(f"Unexpected error during category refinement for script {script_id}, category {category_name}: {e}")
        if db.is_active: 
            try: db.rollback()
            except: pass
        return jsonify({"error": "An unexpected error occurred during category refinement."}), 500
    finally:
        if db:
            db.close()

# --- NEW: Script Refinement Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/refine', methods=['POST'])
def refine_vo_script(script_id: int):
    """Refines all VO script lines for a script using OpenAI based on global,
       category, and line prompts, optionally applying ElevenLabs best practices.
    """
    from backend.app import model_to_dict # Import locally
    data = request.get_json()
    if not data:
        return jsonify({"error": "Missing request body"}), 400
        
    global_prompt = data.get('global_prompt')
    apply_best_practices = data.get('apply_best_practices', False) # NEW: Get flag
    target_model = data.get('model', utils_openai.DEFAULT_REFINEMENT_MODEL)

    # Allow empty global_prompt if apply_best_practices is true
    if not global_prompt and not apply_best_practices:
        return jsonify({"error": "Missing 'global_prompt' or 'apply_best_practices' must be true"}), 400

    db: Session = next(get_db())
    updated_lines_data = []
    errors_occurred = False
    
    # --- NEW: Get ElevenLabs rules if requested --- 
    elevenlabs_rules = None
    if apply_best_practices:
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            prompt_file_dir = os.path.dirname(current_dir)
            rules_path = os.path.join(prompt_file_dir, 'prompts', 'scripthelp.md')
            elevenlabs_rules = _get_elevenlabs_rules(rules_path)
            if not elevenlabs_rules:
                logging.warning(f"Could not load ElevenLabs rules from {rules_path} for script refine. Proceeding without them.")
        except Exception as e:
            logging.exception(f"Error loading ElevenLabs rules for script refine: {e}")
            elevenlabs_rules = None # Ensure it's None if loading failed
    # --- END NEW --- 
    
    try:
        # 1. Get context (includes script/category prompts)
        lines_to_process = utils_voscript.get_script_lines_context(db, script_id)
        
        if not lines_to_process:
            logging.info(f"No lines found for script {script_id}. Nothing to refine.")
            return jsonify({"data": []}), 200 

        logging.info(f"Found {len(lines_to_process)} potential lines to refine for script {script_id}.")

        # 2. Iterate and refine each NON-LOCKED line
        # --- NEW: Create map for quick text lookup --- 
        line_texts_map = {ltx['line_id']: ltx.get('current_text', '') for ltx in lines_to_process}
        
        for line_index, line_context in enumerate(lines_to_process):
            line_id = line_context['line_id']
            
            if line_context.get('is_locked', False):
                logging.info(f"Skipping locked line {line_id} during script refinement.")
                continue 
                
            # --- Construct Prompt (Conditional) --- 
            # Base context (Always included)
            base_context_prompt_parts = [
                f"You are a creative writer for video game voiceovers.",
                f"Character Description:\n{line_context.get('character_description', 'N/A')}\n",
                f"Template Hint: {line_context.get('template_hint', 'N/A')}",
                f"Category: {line_context.get('category_name', 'N/A')}",
                f"Category Instructions: {line_context.get('category_instructions', 'N/A')}",
                f"Line Key: {line_context.get('line_key', 'N/A')}",
                f"Line Hint: {line_context.get('line_template_hint', 'N/A')}\n",
                f"Current Line Text:\n{line_context.get('current_text', '')}"
            ]
            
            # User refinement prompts (Always included, adjusted if empty)
            global_request_prompt = global_prompt or "N/A"
            category_prompt_text = line_context.get('category_refinement_prompt') or "N/A"
            line_feedback_text = line_context.get('latest_feedback') or "N/A"
            user_request_summary = f"Global Script Prompt: {global_request_prompt}\nCategory Prompt: {category_prompt_text}\nLine Feedback/Prompt: {line_feedback_text}"
            user_request_text_for_stage1 = global_prompt if global_prompt else "No specific global refinement request provided."

            # Sibling examples (Calculate only if needed)
            sibling_examples_text_parts = []
            variety_instruction = "" # Initialize as empty
            if len(lines_to_process) > 1:
                sibling_examples = []
                example_count = 0
                for i, sibling_context in enumerate(lines_to_process):
                    if i == line_index or example_count >= 5:
                        continue
                    sibling_key = sibling_context.get('line_key') or f"line_{sibling_context['line_id']}"
                    sibling_text = line_texts_map.get(sibling_context['line_id'], "[ERROR: Text not found]")
                    if sibling_text:
                         sibling_examples.append(f"- {sibling_key}: \"{sibling_text}\"")
                         example_count += 1
                if sibling_examples:
                    sibling_examples_text_parts.append("\n--- Sibling Line Examples (for context and ensuring variety) ---")
                    sibling_examples_text_parts.extend(sibling_examples)
                    # Define variety instruction ONLY if siblings exist
                    variety_instruction = "\n\nIMPORTANT: Ensure the refined output for this specific line is varied and distinct (e.g., in structure, theme, punchline) compared to the Sibling Line Examples provided. Avoid repetitive patterns."

            # ElevenLabs rules (Conditional)
            elevenlabs_rules_parts = []
            if apply_best_practices and elevenlabs_rules:
                elevenlabs_rules_parts.append("\n--- Stage 2: Apply ElevenLabs Best Practices ---")
                elevenlabs_rules_parts.append(f"ElevenLabs Rules:\n{elevenlabs_rules}")

            # Final Instructions (Conditional)
            final_instructions_parts = []
            if apply_best_practices and elevenlabs_rules:
                 # Two-stage instructions
                 final_instructions_parts.append("\n--- Instructions ---")
                 final_instructions_parts.append(f"1. Rewrite the 'Current Line Text' based *only* on the hierarchical user refinement prompts above ({user_request_summary}) and the original context (Character Description, Hints, etc.).")
                 final_instructions_parts.append(f"2. Take the result from step 1 and apply the 'ElevenLabs Rules' to it, adding appropriate formatting like <break time=\"0.5s\"/> tags for pauses, etc.{variety_instruction}")
                 final_instructions_parts.append("3. Output ONLY the final text after applying both stages. Do not include explanations or intermediate steps.\n")
                 final_instructions_parts.append("--- FINAL REFINED AND FORMATTED TEXT ---")
            else:
                 # Single-stage instructions
                 final_instructions_parts.append(f"\n\n--- Prompts (Apply these hierarchically) ---\n{user_request_summary}")
                 final_instructions_parts.append(f"\nRewrite the 'Current Line Text' based on ALL applicable prompts above (Global, Category, Line), while staying consistent with the character description and other hints. Prioritize the most specific prompt if conflicts arise.{variety_instruction}")
                 final_instructions_parts.append("Only output the refined line text, with no extra explanation or preamble. If no change is needed based on the prompts, output the original text exactly.")

            # Combine the parts
            openai_prompt_list = base_context_prompt_parts + sibling_examples_text_parts
            # Add Stage 1 marker only if doing two stages
            if apply_best_practices and elevenlabs_rules:
                openai_prompt_list.append(f"\n\n--- Stage 1: User Refinement Request ---\nUser Request: \"{user_request_text_for_stage1}\"")
            openai_prompt_list.extend(elevenlabs_rules_parts)
            openai_prompt_list.extend(final_instructions_parts)
            
            openai_prompt = "\n".join(openai_prompt_list)
            # --- END REVISED Prompt Construction --- 
            
            logging.debug(f"Sending script-refine prompt to OpenAI for line {line_id} (Apply Rules: {apply_best_practices})...")
            refined_text = utils_openai.call_openai_responses_api(
                prompt=openai_prompt,
                model=target_model
            )

            if refined_text is None:
                logging.error(f"OpenAI script refinement failed for script {script_id}, line {line_id}")
                errors_occurred = True 
                continue 
            
            logging.info(f"Refined text received for line {line_id} via script refine.")

            # Update Database for this line
            new_status = "review" if refined_text != line_context.get('current_text') else line_context.get('status', 'generated')
            updated_line = utils_voscript.update_line_in_db(
                db, line_id, refined_text, new_status, target_model
            )
            
            if updated_line is None:
                logging.error(f"Database update failed after script refinement for script {script_id}, line {line_id}")
                errors_occurred = True
            else:
                updated_lines_data.append(model_to_dict(updated_line)) 

        # 3. Return results
        if errors_occurred:
             status_code = 207 # Multi-Status
             message = "Script refinement completed with some errors."
        else:
             status_code = 200
             message = "Script refinement completed successfully."
             
        return jsonify({"message": message, "data": updated_lines_data}), status_code

    except Exception as e:
        logging.exception(f"Unexpected error during script refinement for script {script_id}: {e}")
        if db.is_active: 
            try: db.rollback()
            except: pass
        return jsonify({"error": "An unexpected error occurred during script refinement."}), 500
    finally:
        if db:
            db.close()

# --- NEW: Line Locking Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/lines/<int:line_id>/toggle-lock', methods=['PATCH'])
def toggle_lock_vo_script_line(script_id: int, line_id: int):
    """Toggles the is_locked status of a specific VO script line."""
    # from backend.app import model_to_dict # Import locally - not needed if manually constructing dict
    db: Session = next(get_db())
    try:
        # Find the specific line, ensuring it belongs to the correct script
        line = db.query(models.VoScriptLine).filter(
            models.VoScriptLine.id == line_id,
            models.VoScriptLine.vo_script_id == script_id
        ).first()

        if not line:
            return jsonify({"error": f"Line not found with ID {line_id} for script {script_id}"}), 404
        
        # Toggle the status
        line.is_locked = not line.is_locked
        new_lock_status = line.is_locked
        
        db.commit()
        db.refresh(line)
        logging.info(f"Toggled lock status for line {line_id} (script {script_id}) to {new_lock_status}")
        
        # Manually construct response dict with specific fields
        response_data = {
            "id": line.id,
            "is_locked": line.is_locked,
            "updated_at": line.updated_at.isoformat() if line.updated_at else None
        }
        return jsonify({"data": response_data}), 200

    except Exception as e:
        db.rollback()
        logging.exception(f"Error toggling lock for line {line_id}, script {script_id}: {e}")
        return jsonify({"error": "Failed to toggle lock status."}), 500
    finally:
        if db:
            db.close()

# --- NEW: Manual Text Update Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/lines/<int:line_id>/update-text', methods=['PATCH'])
def update_vo_script_line_text(script_id: int, line_id: int):
    """Updates the generated_text for a specific line (manual edit)."""
    # from backend.app import model_to_dict # Import locally - No longer needed
    data = request.get_json()
    if not data or 'generated_text' not in data:
        return jsonify({"error": "Missing 'generated_text' in request body"}), 400
    
    new_text = data['generated_text']
    if not isinstance(new_text, str):
         return jsonify({"error": "'generated_text' must be a string"}), 400
         
    db: Session = next(get_db())
    try:
        line = db.query(models.VoScriptLine).filter(
            models.VoScriptLine.id == line_id,
            models.VoScriptLine.vo_script_id == script_id
        ).first()

        if not line:
            return jsonify({"error": f"Line not found with ID {line_id} for script {script_id}"}), 404
        
        original_text = line.generated_text # Capture original text
        original_status = line.status # Capture original status
        
        # --- Add "Before" history entry --- #
        current_history = line.generation_history if isinstance(line.generation_history, list) else []
        pre_history_entry = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "type": "pre_manual_edit", # Indicate state before manual edit
            "text": original_text,
            "model": "user",
            "status_before": original_status # Optional: store previous status
        }
        current_history.append(pre_history_entry)
        # Don't set line.generation_history yet, do it after adding the 'after' entry

        # Update text and set status to 'manual'
        line.generated_text = new_text
        line.status = 'manual' 
        line.latest_feedback = None # Clear feedback on manual edit
        
        # Add "After" history entry
        post_history_entry = {
            "timestamp": datetime.now(timezone.utc).isoformat(), # Use a slightly later timestamp potentially?
            "type": "manual_edit",
            "text": new_text,
            "model": "user"
        }
        current_history.append(post_history_entry)
        line.generation_history = current_history
        # Flag as modified
        flag_modified(line, "generation_history")
        
        db.commit()
        db.refresh(line)
        logging.info(f"Manually updated text for line {line_id} (script {script_id}), logged pre/post history.")
        
        # Manually construct simpler response dict (avoiding potentially unloaded attrs)
        response_data = {
            "id": line.id,
            # "vo_script_id": line.vo_script_id,
            # "template_line_id": line.template_line_id,
            # "category_id": line.category_id, # Removed potentially problematic attribute
            "generated_text": line.generated_text,
            "status": line.status,
            "latest_feedback": line.latest_feedback,
            "generation_history": line.generation_history,
            # "line_key": line.line_key, # Removed potentially problematic attribute
            # "order_index": line.order_index, # Removed potentially problematic attribute
            # "prompt_hint": line.prompt_hint, # Removed potentially problematic attribute
            "is_locked": line.is_locked,
            "created_at": line.created_at.isoformat() if line.created_at else None,
            "updated_at": line.updated_at.isoformat() if line.updated_at else None
        }
        # Return the updated line data using the standard wrapper
        return make_api_response(data=response_data)

    except Exception as e:
        db.rollback()
        logging.exception(f"Error updating text for line {line_id}, script {script_id}: {e}")
        return make_api_response(error="Failed to update line text.", status_code=500) # Use standard wrapper
    finally:
        if db:
            db.close()
            
# --- NEW: Delete Line Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/lines/<int:line_id>', methods=['DELETE'])
def delete_vo_script_line(script_id: int, line_id: int):
    """Deletes a specific VO script line."""
    db: Session = next(get_db())
    try:
        line = db.query(models.VoScriptLine).filter(
            models.VoScriptLine.id == line_id,
            models.VoScriptLine.vo_script_id == script_id
        ).first()

        if not line:
            return jsonify({"error": f"Line not found with ID {line_id} for script {script_id}"}), 404
            
        db.delete(line)
        db.commit()
        logging.info(f"Deleted line {line_id} from script {script_id}")
        
        return jsonify({"message": "Line deleted successfully"}), 200

    except Exception as e:
        db.rollback()
        logging.exception(f"Error deleting line {line_id}, script {script_id}: {e}")
        return jsonify({"error": "Failed to delete line."}), 500
    finally:
        if db:
            db.close()

# --- NEW: Add Line Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/lines', methods=['POST'])
def add_vo_script_line(script_id: int):
    """Adds a new custom line to a VO script."""
    # from backend.app import model_to_dict # Import locally - Not needed now
    data = request.get_json()
    if not data:
        return jsonify({"error": "Missing request body"}), 400
    
    # Validate required fields
    line_key = data.get('line_key')
    category_name = data.get('category_name')
    order_index = data.get('order_index')
    if not line_key:
        return jsonify({"error": "Missing 'line_key' in request body"}), 400
    if not category_name:
        return jsonify({"error": "Missing 'category_name' in request body"}), 400
    if order_index is None: # Allow 0
        return jsonify({"error": "Missing 'order_index' in request body"}), 400
        
    # Optional fields
    initial_text = data.get('initial_text')
    prompt_hint = data.get('prompt_hint')
    
    db: Session = next(get_db())
    try:
        # 1. Find parent script
        script = db.query(models.VoScript).get(script_id)
        if not script:
            return jsonify({"error": f"Script not found with ID {script_id}"}), 404
        if not script.template_id:
            # Should not happen based on current logic, but good check
            return jsonify({"error": f"Script {script_id} does not have an associated template.", "detail": "Cannot determine category without a template."}), 400
            
        # 2. Find the category ID based on name and script's template ID
        category = db.query(models.VoScriptTemplateCategory).filter(
            models.VoScriptTemplateCategory.template_id == script.template_id,
            models.VoScriptTemplateCategory.name == category_name,
            models.VoScriptTemplateCategory.is_deleted == False # Ensure category not deleted
        ).first()
        
        if not category:
            return jsonify({"error": f"Category '{category_name}' not found for the template associated with script {script_id}"}), 404

        # 3. Create the new line
        new_line = models.VoScriptLine(
            vo_script_id=script_id,
            template_line_id=None, 
            generated_text=initial_text, 
            status='manual' if initial_text else 'pending',
            is_locked=False
        )
        # Set attributes after creation (Necessary due to model constructor)
        new_line.category_id = category.id
        new_line.line_key = line_key
        new_line.order_index = order_index
        new_line.prompt_hint = prompt_hint
        
        db.add(new_line)
        db.flush() # Explicitly flush to send INSERT to DB before commit
        db.commit()
        db.refresh(new_line)
        logging.info(f"Added new custom line (key: {line_key}) with ID {new_line.id} to script {script_id} under category {category_name} (ID: {category.id})")
        
        # Manually construct response including category_id
        response_data = {
            "id": new_line.id,
            "vo_script_id": new_line.vo_script_id,
            "template_line_id": new_line.template_line_id, # Will be null
            "category_id": new_line.category_id, # Include this!
            "generated_text": new_line.generated_text,
            "status": new_line.status,
            "latest_feedback": new_line.latest_feedback,
            "generation_history": new_line.generation_history,
            "line_key": new_line.line_key,
            "order_index": new_line.order_index,
            "prompt_hint": new_line.prompt_hint,
            "is_locked": new_line.is_locked,
            "created_at": new_line.created_at.isoformat() if new_line.created_at else None,
            "updated_at": new_line.updated_at.isoformat() if new_line.updated_at else None
        }
        return jsonify({"data": response_data}), 201 # Use jsonify directly

    except IntegrityError as e:
        db.rollback()
        # Potential issue: duplicate line_key within the script? (Need constraint?)
        logging.exception(f"Database integrity error adding line to script {script_id}: {e}")
        return jsonify({"error": "Database error adding line. Potential duplicate key?"}), 500
    except Exception as e:
        db.rollback()
        logging.exception(f"Error adding line to script {script_id}: {e}")
        return jsonify({"error": "Failed to add line."}), 500
    finally:
        if db:
            db.close()

# --- NEW: Single Line Generation Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/lines/<int:line_id>/generate', methods=['POST'])
def generate_vo_script_line(script_id: int, line_id: int):
    """Generates text for a single VO script line using OpenAI.
       Typically used for lines in 'pending' status.
    """
    from backend.app import model_to_dict # Import locally
    
    # Optional: Get model override from request? Not strictly needed for generate draft.
    # data = request.get_json() 
    # target_model = data.get('model', utils_openai.DEFAULT_GENERATION_MODEL) if data else utils_openai.DEFAULT_GENERATION_MODEL
    target_model = utils_openai.DEFAULT_REFINEMENT_MODEL # Use refinement model for now

    db: Session = next(get_db())
    try:
        # 1. Get context for the line
        line_context = utils_voscript.get_line_context(db, line_id)
        if not line_context:
            logging.warning(f"Generate line request: Context not found for script {script_id}, line {line_id}")
            return make_api_response(error=f"Line context not found for line_id {line_id}", status_code=404)
        
        # 2. Get sibling lines in the same category for context and to ensure variety
        category_name = line_context.get('category_name')
        sibling_lines = []
        
        if category_name:
            sibling_lines = utils_voscript.get_sibling_lines_in_category(
                db, script_id, line_id, category_name, limit=10
            )
            logging.info(f"Found {len(sibling_lines)} sibling lines in category '{category_name}' for context")
        
        # 3. Construct prompt with sibling context for variety
        prompt_parts = [
            f"You are a creative writer for video game voiceovers.",
            f"Character Description:\n{line_context.get('character_description', 'N/A')}\n",
            f"Template Hint: {line_context.get('template_hint', 'N/A')}",
            f"Category: {line_context.get('category_name', 'N/A')}",
            f"Category Instructions: {line_context.get('category_instructions', 'N/A')}",
            f"Line Key: {line_context.get('line_key', 'N/A')}",
            f"Line Hint: {line_context.get('line_template_hint', 'N/A')}\n"
        ]
        
        # Add sibling lines if available
        if sibling_lines:
            prompt_parts.append("\n--- Existing Lines in This Category (For Context) ---")
            for sibling in sibling_lines:
                prompt_parts.append(f"- {sibling['line_key']}: \"{sibling['text']}\"")
                
            # Add strong variety instructions
            prompt_parts.append("\n--- IMPORTANT: VARIETY REQUIREMENTS ---")
            prompt_parts.append("Your task is to write a new line that is DISTINCTLY DIFFERENT from all existing lines.")
            prompt_parts.append("Requirements:")
            prompt_parts.append("1. DO NOT use the same sentence structure as any existing line")
            prompt_parts.append("2. DO NOT repeat phrases, jokes, or puns that appear in existing lines")
            prompt_parts.append("3. DO NOT start with the same words/phrases as other lines (e.g., if others start with 'Get ready', use a different opening)")
            prompt_parts.append("4. USE UNIQUE vocabulary and expressions not present in other lines")
            prompt_parts.append("5. VARY the emotional tone or attitude (e.g., if others are boastful, make this one mysterious)")
            prompt_parts.append("6. If multiple lines use similar themes (like food/cooking references), explore a COMPLETELY DIFFERENT theme while staying true to the character")
        
        # Final generation instructions
        prompt_parts.append("\n--- GENERATION TASK ---")
        prompt_parts.append("Write a single voiceover line based on the character description, category, line key, and hints provided.")
        if sibling_lines:
            prompt_parts.append("Make this line DISTINCTLY DIFFERENT from all existing lines shown above while maintaining the character's personality.")
        prompt_parts.append("Only output the voiceover line text, with no explanation or preamble.")
        
        openai_prompt = "\n".join(prompt_parts)
        logging.info(f"Sending enhanced generation prompt to OpenAI for line {line_id}. Prompt includes {len(sibling_lines)} sibling examples.")
        logging.debug(f"Prompt start: {openai_prompt[:200]}...")

        # 4. Call OpenAI Responses API
        generated_text = utils_openai.call_openai_responses_api(
            prompt=openai_prompt,
            model=target_model
        )
        
        if generated_text is None:
            logging.error(f"OpenAI generation failed for script {script_id}, line {line_id}")
            return make_api_response(error="OpenAI generation failed. Check logs.", status_code=500)
            
        logging.info(f"Generated text received for line {line_id}: '{generated_text[:100]}...'")

        # 5. Update Database
        new_status = "generated" # Set status after successful generation
        updated_line = utils_voscript.update_line_in_db(
            db, 
            line_id, 
            generated_text, 
            new_status, 
            target_model # Log which model did the generation
        )
        
        if updated_line is None:
            logging.error(f"Database update failed after generation for script {script_id}, line {line_id}")
            return make_api_response(error="Database update failed after generation.", status_code=500)

        # 6. Return updated line data
        return make_api_response(data=model_to_dict(updated_line))

    except Exception as e:
        logging.exception(f"Unexpected error during line generation for script {script_id}, line {line_id}: {e}")
        if db.is_active:
             try: db.rollback()
             except: pass
        return make_api_response(error="An unexpected error occurred during generation.", status_code=500)
    finally:
        if db:
            db.close()

# --- NEW: Accept Line Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/lines/<int:line_id>/accept', methods=['PATCH'])
def accept_vo_script_line(script_id: int, line_id: int):
    """Marks a VO script line with status 'review' as 'generated'."""
    from backend.app import model_to_dict # Import locally
    db: Session = next(get_db())
    try:
        line = db.query(models.VoScriptLine).filter(
            models.VoScriptLine.id == line_id,
            models.VoScriptLine.vo_script_id == script_id
        ).first()

        if not line:
            return make_api_response(error=f"Line not found with ID {line_id} for script {script_id}", status_code=404)
            
        if line.status != 'review':
            # Optionally return an error or just return the current state if not in review
            logging.warning(f"Attempted to accept line {line_id} which is not in 'review' status (current: {line.status}). Returning current state.")
            return make_api_response(data=model_to_dict(line)) # Return current state
            # OR return make_api_response(error="Line is not in 'review' status.", status_code=400)
        
        # Update status
        line.status = 'generated' 
        
        db.commit()
        db.refresh(line)
        logging.info(f"Accepted line {line_id} (script {script_id}), status set to {line.status}.")
        
        # Return updated line data
        return make_api_response(data=model_to_dict(line))

    except Exception as e:
        db.rollback()
        logging.exception(f"Error accepting line {line_id}, script {script_id}: {e}")
        return make_api_response(error="Failed to accept line.", status_code=500)
    finally:
        if db:
            db.close()

# --- NEW: Batch Category Generation Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/categories/<category_name>/generate-batch', methods=['POST'])
def generate_category_lines_batch(script_id: int, category_name: str):
    """Generates text for all pending lines in a category together, ensuring variety.
       Uses a two-pass approach:
       1. First pass: OpenAI prompt to create varied lines in a batch
       2. Second pass: If needed, iteratively refine any lines that are too similar
    """
    from backend.app import model_to_dict # Import locally
    
    # Get optional model override
    data = request.get_json() or {}
    target_model = data.get('model', utils_openai.DEFAULT_REFINEMENT_MODEL)
    
    db: Session = next(get_db())
    updated_lines_data = []
    errors_occurred = False
    
    try:
        # 1. Get all pending lines in the category
        lines_to_process = utils_voscript.get_category_lines_context(db, script_id, category_name)
        
        if not lines_to_process:
            logging.info(f"No lines found for category '{category_name}' in script {script_id}. Nothing to generate.")
            return make_api_response(data={"message": "No lines found to generate", "data": []})
            
        # DEBUG: Log each line's status, id, and key to see what we have
        for line in lines_to_process:
            logging.info(f"Found line: ID={line.get('line_id')}, key={line.get('line_key')}, status={line.get('status')}, has_text={line.get('current_text') is not None}")
        
        # Filter to only include pending lines or empty lines
        pending_lines = []
        for line in lines_to_process:
            # Check for explicitly 'pending' status
            if line.get('status') == 'pending':
                pending_lines.append(line)
                logging.info(f"Added pending line ID {line.get('line_id')} with key {line.get('line_key')} to generation queue")
            # Also check for any empty lines regardless of status
            elif not line.get('current_text') and not line.get('is_locked', False):
                pending_lines.append(line)
                logging.info(f"Added empty line ID {line.get('line_id')} with key {line.get('line_key')} to generation queue")
            # Special check for DIRECTED_TAUNT lines
            elif line.get('line_key', '').startswith('DIRECTED_TAUNT_') and not line.get('is_locked', False):
                pending_lines.append(line)
                logging.info(f"Added directed taunt line ID {line.get('line_id')} with key {line.get('line_key')} to generation queue")
        
        if not pending_lines:
            logging.info(f"No pending lines found for category '{category_name}' in script {script_id}. Nothing to generate.")
            return make_api_response(data={"message": "No pending lines found to generate", "data": []})
            
        logging.info(f"Found {len(pending_lines)} pending lines to generate for category '{category_name}' in script {script_id}.")

        # 2. Find any existing generated lines in the same category (for context and variety)
        existing_lines = [line for line in lines_to_process if line.get('status') not in ['pending', None] and 
                           line.get('current_text') and
                           line.get('line_id') not in [pl.get('line_id') for pl in pending_lines]]
        
        # 3. If we have 10 or fewer pending lines, use batch generation approach
        if len(pending_lines) <= 10:
            # APPROACH 1: Batch Generation (Ideal for smaller batches)
            updated_lines_data = _generate_lines_batch(db, script_id, pending_lines, existing_lines, target_model)
        else:
            # APPROACH 2: Split into smaller batches (For larger sets)
            logging.info(f"Large batch of {len(pending_lines)} lines detected. Splitting into smaller batches.")
            
            # Process in batches of 8 lines
            batch_size = 8
            for i in range(0, len(pending_lines), batch_size):
                batch = pending_lines[i:i+batch_size]
                batch_existing = existing_lines + updated_lines_data  # Include already generated lines as context
                
                logging.info(f"Processing batch {i//batch_size + 1} with {len(batch)} lines")
                batch_results = _generate_lines_batch(db, script_id, batch, batch_existing, target_model)
                
                if batch_results:
                    updated_lines_data.extend(batch_results)
                    # Update existing_lines for next iteration
                    existing_lines = [line for line in existing_lines if line.get('line_id') not in [l.get('id') for l in batch_results]]
                else:
                    errors_occurred = True
                
        # 4. Return results
        if errors_occurred:
            status_code = 207  # Multi-Status
            message = "Category generation completed with some errors."
        else:
            status_code = 200
            message = "Category generation completed successfully."
            
        return make_api_response(
            data={
                "message": message, 
                "data": updated_lines_data
            },
            status_code=status_code
        )
        
    except Exception as e:
        logging.exception(f"Unexpected error during category batch generation for script {script_id}, category {category_name}: {e}")
        if db and db.is_active:
            try: db.rollback()
            except: pass
        return make_api_response(
            error="An unexpected error occurred during batch generation.", 
            status_code=500
        )
    finally:
        if db and db.is_active:
            db.close()

# --- Helper for Batch Generation --- #
# @limits.limit("10 per minute") # Rate limit if needed
# @retry(stop=stop_after_attempt(3), wait=wait_fixed(2)) # Retry logic if needed
def _generate_lines_batch(db: Session, script_id: int, pending_lines: list, existing_lines: list | None, target_model: str) -> list:
    """Helper function to generate multiple lines in a batch with variety.
       Fetches its own limited context based on order_index of pending_lines.

    Args:
        db: Database session.
        script_id: Parent VO Script ID.
        pending_lines: List of line context dicts for lines needing generation in this batch.
        existing_lines: (No longer used effectively - kept for signature compatibility but ignored)
        target_model: OpenAI model name.

    Returns:
        List of updated line context dicts with 'generated_text'.
    """
    logging.info(f"_generate_lines_batch called for script {script_id} with {len(pending_lines)} pending lines. Model: {target_model}")
    if not pending_lines:
        return []

    # Ensure we have a valid model - use default if none provided
    if not target_model:
        target_model = utils_openai.DEFAULT_GENERATION_MODEL
        logging.info(f"No target model specified, using default model: {target_model}")
    
    # Use 'client' directly from utils_openai
    openai_client = utils_openai.client
    updated_lines = [] # Store results {line_id: id, generated_text: text}
    
    # Ensure we only process lines that are actually pending or empty
    lines_to_generate = []
    for line in pending_lines:
        # Check if text is missing OR status is pending (even if text exists, e.g., from previous failed attempt)
        if not line.get('current_text') or line.get('status') == 'pending':
            lines_to_generate.append(line)
        else:
            # If line somehow got here but already has text and isn't pending,
            # add it to updated_lines as is, so it's included in the return count
            updated_lines.append({
                "line_id": line.get('line_id'),
                "generated_text": line.get('current_text')
            })
            logging.info(f"Skipping line ID {line.get('line_id')} in batch generation as it already has text and status is not pending.")
            
    # If all lines already have text, return what we have
    if not lines_to_generate:
        logging.info(f"All lines in batch already have text, skipping batch generation")
        return updated_lines
        
    # Continue with generation for remaining lines
    pending_lines = lines_to_generate # Overwrite with only the lines needing generation
    
    # 1. Get common context from first line (should be same for all lines in category)
    first_pending_line = pending_lines[0]
    char_desc = first_pending_line.get('character_description', 'N/A')
    template_hint = first_pending_line.get('template_hint', 'N/A')
    category_name = first_pending_line.get('category_name', 'N/A')
    first_line_id = first_pending_line.get('line_id') # Needed for category instruction lookup
    
    # --- Fetch category instructions --- 
    category_instructions = "N/A" # Default
    parent_script_template_id = None
    # Get template ID from the script associated with the first line
    if first_line_id:
         first_line_obj = db.query(models.VoScriptLine).options(joinedload(models.VoScriptLine.vo_script)).get(first_line_id)
         if first_line_obj and first_line_obj.vo_script:
             parent_script_template_id = first_line_obj.vo_script.template_id
    
    if parent_script_template_id and category_name != "Uncategorized":
         category = db.query(models.VoScriptTemplateCategory).filter(
             models.VoScriptTemplateCategory.template_id == parent_script_template_id,
             models.VoScriptTemplateCategory.name == category_name
         ).first()
         if category and category.prompt_instructions:
             category_instructions = category.prompt_instructions
             logging.info(f"Using category instructions for '{category_name}' from DB.")
         else:
             logging.warning(f"Could not find category instructions for '{category_name}' in DB, using default.")
    else:
         logging.warning(f"Could not determine template ID or category name ('{category_name}') to fetch instructions, using default.")

    # --- NEW: Fetch Limited Context (Nearby Lines) ---
    context_lines = []
    context_lines_count = 5 # Fetch 5 before and 5 after
    try:
        # Get order indices of the current batch
        batch_order_indices = sorted([l.get('order_index') for l in pending_lines if l.get('order_index') is not None])
        
        if batch_order_indices:
            min_order = batch_order_indices[0]
            max_order = batch_order_indices[-1]
            category_id = first_pending_line.get('category_id') # Get category ID from first line

            if category_id:
                # Fetch preceding lines
                preceding_lines_q = db.query(
                        models.VoScriptLine.line_key, 
                        models.VoScriptLine.generated_text, 
                        models.VoScriptTemplateLine.order_index
                    ).join(
                        models.VoScriptLine.template_line
                    ).filter(
                        models.VoScriptLine.vo_script_id == script_id,
                        models.VoScriptLine.category_id == category_id,
                        models.VoScriptTemplateLine.order_index < min_order,
                        models.VoScriptLine.generated_text.isnot(None),
                        models.VoScriptLine.generated_text != ''
                    ).order_by(
                        models.VoScriptTemplateLine.order_index.desc()
                    ).limit(context_lines_count)

                # Fetch succeeding lines
                succeeding_lines_q = db.query(
                        models.VoScriptLine.line_key, 
                        models.VoScriptLine.generated_text, 
                        models.VoScriptTemplateLine.order_index
                    ).join(
                        models.VoScriptLine.template_line
                    ).filter(
                        models.VoScriptLine.vo_script_id == script_id,
                        models.VoScriptLine.category_id == category_id,
                        models.VoScriptTemplateLine.order_index > max_order,
                        models.VoScriptLine.generated_text.isnot(None),
                        models.VoScriptLine.generated_text != ''
                    ).order_by(
                        models.VoScriptTemplateLine.order_index.asc()
                    ).limit(context_lines_count)
                
                context_lines_db = preceding_lines_q.all() + succeeding_lines_q.all()
                
                # Convert to dictionary format expected by prompt builder
                context_lines = [
                    {'line_key': l.line_key, 'current_text': l.generated_text} 
                    for l in context_lines_db if l.generated_text # Ensure text exists
                ]
                logging.info(f"Fetched {len(context_lines)} nearby lines for context.")
            else:
                 logging.warning(f"Could not determine category_id for context fetching.")
        else:
            logging.warning("No valid order_index found in batch, cannot fetch context lines.")
    except Exception as context_exc:
        logging.error(f"Error fetching context lines: {context_exc}")
    # --- END NEW CONTEXT FETCH ---

    # 2. Build the batch prompt
    prompt_parts = [
        f"You are a creative writer for video game voiceovers.",
        f"Character Description:\\n{char_desc}\\n",
        f"Template Hint: {template_hint}",
        f"Category: {category_name}",
        f"Category Instructions: {category_instructions}\\n"
    ]
    
    # Add existing lines for context (use the limited context_lines now)
    if context_lines:
        prompt_parts.append("\n--- Nearby Lines in This Category (For Context) ---")
        for line in context_lines:
            # Use line_key directly from the query result
            key = line.get('line_key') or f"unknown_line" # Fallback needed? Query includes key.
            text = line.get('current_text', '')
            if key and text: # Ensure both key and text exist
                prompt_parts.append(f'- {key}: "{text}"') # Use single quotes for f-string
    
    # Add variety requirements
    prompt_parts.append("\n--- IMPORTANT: VARIETY REQUIREMENTS ---")
    prompt_parts.append("Your task is to write NEW, VARIED lines that are DISTINCTLY DIFFERENT from each other and from the nearby context lines.")
    prompt_parts.append("Requirements:")
    prompt_parts.append("- Ensure each generated line is unique.")
    prompt_parts.append("- Avoid repetition in phrasing, sentence structure, and core ideas compared to context lines AND other lines in this batch.")
    prompt_parts.append("- Maintain the character's voice and tone.")
    prompt_parts.append("- Fulfill the specific request/hint for each line key.")

    prompt_parts.append("\n--- Lines to Generate (Provide JSON output) ---")
    prompt_parts.append("Generate text for the following line keys. Provide ONLY a valid JSON object where keys are the line keys and values are the generated text strings:")

    # Add lines to be generated with their hints
    lines_to_request_json = {}
    for line in pending_lines:
        key = line.get('line_key') # Use the key directly from context
        if not key: key = f"line_{line.get('line_id')}" # Fallback if key missing
        hint = line.get('prompt_hint', line.get('template_prompt_hint', 'Generate appropriate text.'))
        lines_to_request_json[key] = hint
        # Also add to prompt for clarity for the LLM
        prompt_parts.append(f"- {key}: (Hint: {hint})")
        
    prompt_parts.append("\nJSON Output:") # Cue for the LLM

    full_prompt = "\n".join(prompt_parts)
    # logging.debug(f"Full Batch Prompt:\n{full_prompt}") # DEBUG: Careful logging large prompts

    try:
        logging.info(f"Sending batch generation request to OpenAI model {target_model} for {len(pending_lines)} lines.")
        
        # Log the pending lines and their keys for debugging
        logging.info(f"Pending lines data: {[{'id': l.get('id', l.get('line_id')), 'line_key': l.get('line_key')} for l in pending_lines]}")
        
        response = utils_openai.client.chat.completions.create(
            model=target_model,
            messages=[
                {"role": "system", "content": "You are a helpful assistant providing JSON output."},
                {"role": "user", "content": full_prompt}
            ],
            response_format={"type": "json_object"},
            temperature=0.8, # Increase variety slightly?
            # max_tokens=?? # Set appropriate limit if needed
        )
        
        generated_json_str = response.choices[0].message.content
        logging.info(f"Received response from OpenAI.")
        # logging.debug(f"Generated JSON String: {generated_json_str}") # DEBUG
        
        generated_data = json.loads(generated_json_str)
        
        # Log the JSON keys received from OpenAI
        logging.info(f"Keys received from OpenAI: {list(generated_data.keys())}")
        
        # Map results back to line IDs
        key_to_id_map = {}
        for l in pending_lines:
            # Get the line ID (could be in 'id' or 'line_id' field)
            line_id = l.get('line_id')
            if line_id is None:
                line_id = l.get('id')  # Try alternate field name
            
            # Get the line key (with fallback to line_{id})
            line_key = l.get('line_key')
            if not line_key and line_id is not None:
                line_key = f"line_{line_id}"
                
            # Only add to map if we have both key and id
            if line_key and line_id is not None:
                key_to_id_map[line_key] = line_id
            else:
                logging.warning(f"Could not create key mapping for line: {l} - missing key or ID")
        
        # Log the mapping dictionary for debugging
        logging.info(f"key_to_id_map: {key_to_id_map}")
        
        generated_count = 0
        for key, text in generated_data.items():
            line_id = key_to_id_map.get(key)
            if line_id and isinstance(text, str) and text.strip():
                updated_lines.append({
                    "line_id": line_id,  # Consistently use line_id as the field name
                    "generated_text": text.strip()
                })
                generated_count += 1
                logging.info(f"Successfully mapped key '{key}' to line ID {line_id}")
            else:
                 logging.warning(f"Could not map generated key '{key}' back to line ID or received invalid text. Valid keys: {list(key_to_id_map.keys())}")
                 
        logging.info(f"Successfully processed response, generated text for {generated_count}/{len(pending_lines)} requested lines.")

    except json.JSONDecodeError as json_err:
        logging.error(f"Failed to parse JSON response from OpenAI: {json_err}")
        logging.error(f"Received content: {generated_json_str[:500]}...") # Log beginning of invalid response
        # How to handle? Return empty? Raise error?
        # For now, return whatever was processed before error + any lines skipped initially
        # Task logic needs to handle potential partial success/failure.
        raise Exception(f"Failed to parse OpenAI JSON response: {json_err}") from json_err
    except Exception as e:
        logging.exception(f"Error during OpenAI call in _generate_lines_batch: {e}")
        # Re-raise to be caught by the task
        raise Exception(f"OpenAI API call failed: {e}") from e

    return updated_lines

# Add a new button in "Generate Draft" menu to generate lines by category

# Other VoScript endpoints (GET lines, POST feedback) can be added later

# Other VoScript CRUD endpoints will be added here... 

# --- NEW: Batch Category Generation Task Trigger --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/categories/<category_name>/generate-batch-task', methods=['POST'])
def trigger_category_batch_generation(script_id: int, category_name: str):
    """Triggers the batch generation of all pending lines in a category via Celery task."""
    # Get optional parameters from request body
    data = request.get_json() or {}
    target_model = data.get('model')  # Optional model override
    
    db: Session = None
    db_job = None
    try:
        db = next(get_db())
        # Verify script exists
        script = db.query(models.VoScript).get(script_id)
        if not script:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)
            
        # Create Job record
        job_params = {
            "task_type": "generate_category",
            "category_name": category_name
        }
        if target_model:
            job_params["model"] = target_model
             
        db_job = models.GenerationJob(
            status="PENDING",
            job_type="script_creation", 
            target_batch_id=str(script_id),
            parameters_json=json.dumps(job_params)
        )
        db.add(db_job)
        db.commit()
        db.refresh(db_job)
        db_job_id = db_job.id
        logging.info(f"Created Category Batch Generation Job DB ID: {db_job_id} for VO Script ID {script_id}, Category '{category_name}'")

        # Enqueue Celery task
        task = tasks.generate_category_lines.delay(
            db_job_id,
            script_id,
            category_name,
            target_model
        )
        logging.info(f"Enqueued category batch generation task: Celery ID {task.id}, DB Job ID {db_job_id}")
        
        # Link Celery task ID to Job record
        db_job.celery_task_id = task.id
        db.commit()
        
        return make_api_response(data={'job_id': db_job_id, 'task_id': task.id}, status_code=202)
        
    except Exception as e:
        logging.exception(f"Error submitting category batch generation job for script {script_id}, category '{category_name}': {e}")
        if db_job and db_job.id and db.is_active: # Mark DB job as failed if possible
            try: 
                db_job.status = "SUBMIT_FAILED"
                db_job.result_message = f"Enqueue failed: {e}"
                db.commit()
            except: 
                db.rollback()
        elif db and db.is_active:
             db.rollback()
        return make_api_response(error="Failed to start category batch generation task", status_code=500)
    finally:
        if db and db.is_active: db.close()

# --- NEW: Category Variety Analysis Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/categories/<category_name>/analyze-variety', methods=['GET'])
def analyze_category_variety_endpoint(script_id: int, category_name: str):
    """Analyzes the variety of lines in a category to identify repetition and similarity issues."""
    db: Session = next(get_db())
    try:
        # Verify script exists
        script = db.query(models.VoScript).get(script_id)
        if not script:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)
        
        # Verify category exists for this script
        if script.template_id:
            category = db.query(models.VoScriptTemplateCategory).filter(
                models.VoScriptTemplateCategory.template_id == script.template_id,
                models.VoScriptTemplateCategory.name == category_name,
                models.VoScriptTemplateCategory.is_deleted == False
            ).first()
            
            if not category:
                return make_api_response(error=f"Category '{category_name}' not found for script {script_id}", status_code=404)
        else:
            # If script has no template_id, check if any lines have this category name
            line_with_category = db.query(models.VoScriptLine).join(
                models.VoScriptLine.template_line
            ).join(
                models.VoScriptTemplateLine.category
            ).filter(
                models.VoScriptLine.vo_script_id == script_id,
                models.VoScriptTemplateCategory.name == category_name
            ).first()
            
            if not line_with_category:
                return make_api_response(error=f"Category '{category_name}' not found for script {script_id}", status_code=404)
        
        # Perform variety analysis
        analysis_results = utils_voscript.analyze_category_variety(db, script_id, category_name)
        
        # Return the results
        return make_api_response(data=analysis_results)
    
    except Exception as e:
        logging.exception(f"Error analyzing variety for script {script_id}, category '{category_name}': {e}")
        return make_api_response(error=f"Failed to analyze category variety: {str(e)}", status_code=500)
    finally:
        if db:
            db.close()

# --- NEW: Excel Download Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/download-excel', methods=['GET'])
def download_vo_script_excel(script_id: int):
    """Generates and returns an Excel file for a VO script."""
    db: Session = None
    try:
        db = next(get_db())
        # Fetch script details, similar to get_vo_script but leaner if possible
        script = db.query(models.VoScript).options(
            joinedload(models.VoScript.template), # Load template info
            selectinload(models.VoScript.lines).selectinload(models.VoScriptLine.template_line).selectinload(models.VoScriptTemplateLine.category) # Load lines -> template line -> category
        ).get(script_id)
        
        if script is None:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)
            
        # Create Excel Workbook and Sheet
        wb = Workbook()
        ws = wb.active
        ws.title = script.name[:30] # Use script name for sheet title (Excel limit ~31 chars)

        # --- Define Styles ---
        header_font = Font(name='Calibri', size=14, bold=True)
        category_font = Font(name='Calibri', size=12, bold=True)
        category_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid") # Light green fill
        wrapped_alignment = Alignment(wrap_text=True, vertical='top')

        # --- Populate Header Info ---
        # Script Name (A1)
        ws['A1'] = script.name
        ws['A1'].font = header_font
        
        # Character Description (A2, merged)
        ws['A2'] = script.character_description
        # Merge cells A2 to E2 (or adjust as needed)
        merge_range = 'A2:E2' 
        ws.merge_cells(merge_range) 
        ws['A2'].alignment = wrapped_alignment
        # Estimate row height based on text length (basic estimation)
        desc_len = len(script.character_description or "")
        estimated_lines = max(1, (desc_len // 80) + 1) # Rough guess: 80 chars per line
        ws.row_dimensions[2].height = estimated_lines * 15 # 15 points per line height

        # --- Organize Lines by Category ---
        lines_by_category = {}
        for line in script.lines:
            # Determine category (similar logic to get_vo_script)
            category_id = getattr(line, 'category_id', None)
            category_name = "Uncategorized"
            category_obj = None
            if category_id: # Line has direct category_id (custom line)
                category_obj = db.query(models.VoScriptTemplateCategory).get(category_id)
            elif line.template_line and line.template_line.category: # Line linked via template
                 category_obj = line.template_line.category
            
            if category_obj:
                category_name = category_obj.name
                category_id = category_obj.id # Ensure we have the ID
            
            if category_id not in lines_by_category:
                 lines_by_category[category_id] = {'name': category_name, 'lines': []}
            
            # Simplified line data for Excel
            line_data = {
                'key': getattr(line, 'line_key', line.template_line.line_key if line.template_line else f'line_{line.id}'),
                'text': line.generated_text or "",
                'order': getattr(line, 'order_index', line.template_line.order_index if line.template_line else float('inf'))
            }
            lines_by_category[category_id]['lines'].append(line_data)

        # Sort lines within each category, handling None values
        for cat_id in lines_by_category:
            lines_by_category[cat_id]['lines'].sort(key=lambda x: x['order'] if x['order'] is not None else float('inf'))

        # --- Populate Categories and Lines ---
        current_row = 4 # Start after header and a blank row
        
        # Sort categories by name (optional, but makes sense)
        sorted_category_ids = sorted(lines_by_category.keys(), key=lambda cid: lines_by_category[cid]['name'])
        
        for category_id in sorted_category_ids:
            category_data = lines_by_category[category_id]
            category_name = category_data['name']
            
            # Category Header Row
            cat_cell = ws[f'A{current_row}']
            cat_cell.value = category_name
            cat_cell.font = category_font
            cat_cell.fill = category_fill
            # Merge category header across a few columns for visibility
            ws.merge_cells(f'A{current_row}:C{current_row}') 
            current_row += 1
            
            # Line Rows
            for line in category_data['lines']:
                key_cell = ws[f'B{current_row}']
                text_cell = ws[f'C{current_row}']
                
                key_cell.value = line['key']
                text_cell.value = line['text']
                text_cell.alignment = wrapped_alignment
                
                current_row += 1
                
            # Add a blank row between categories
            current_row += 1

        # --- Set Column Widths ---
        ws.column_dimensions['A'].width = 30 # Category name
        ws.column_dimensions['B'].width = 25 # Line Key
        ws.column_dimensions['C'].width = 80 # Line Text

        # --- Prepare Response ---
        # Save workbook to a BytesIO buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0) # Rewind buffer to the beginning

        # Sanitize filename
        sanitized_name = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in script.name)
        # Strip any trailing underscores that might result from sanitization
        sanitized_name = sanitized_name.rstrip('_') 
        # Also strip leading/trailing whitespace just in case
        sanitized_name = sanitized_name.strip()
        # Ensure filename isn't empty after stripping
        if not sanitized_name:
            sanitized_name = "vo_script"
        filename = f"{sanitized_name}.xlsx"
        
        logging.info(f"Final calculated filename for download: '{filename}'") # ADD THIS LINE
        logging.info(f"Generated Excel file for VO Script ID {script_id}")
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.exception(f"Error generating Excel for VO script {script_id}: {e}")
        # Ensure DB session is closed if active
        if db and db.is_active: 
            try: db.rollback() 
            except: pass
        return make_api_response(error="Failed to generate Excel file", status_code=500)
    finally:
        if db and db.is_active:
            db.close()

# --- NEW: Instantiate Target Lines Endpoint --- #
@vo_script_bp.route('/vo-scripts/<int:script_id>/instantiate-lines', methods=['POST'])
def instantiate_target_lines(script_id: int):
    """Creates multiple new 'pending' VO script lines based on a list of targets
       and associates them with a specific category.
    """
    data = request.get_json()
    if not data:
        return make_api_response(error="Missing request body", status_code=400)

    # --- Get Required Inputs --- #
    template_category_id = data.get('template_category_id')
    target_names = data.get('target_names') # Expected to be a list of strings

    # --- Get Optional Inputs --- #
    line_key_prefix = data.get('line_key_prefix', 'TARGETED_LINE_') # Default prefix
    prompt_hint_template = data.get('prompt_hint_template', 'Line targeting {TargetName}') # Default hint template
    # Placeholder for target name in the template
    target_placeholder = data.get('target_placeholder', '{TargetName}') 

    # --- Validate Inputs --- #
    if not template_category_id or not isinstance(template_category_id, int):
        return make_api_response(error="Missing or invalid 'template_category_id' (must be integer)", status_code=400)
    if not target_names or not isinstance(target_names, list) or not all(isinstance(t, str) for t in target_names):
        return make_api_response(error="Missing or invalid 'target_names' (must be a list of strings)", status_code=400)
    if not isinstance(line_key_prefix, str):
        return make_api_response(error="Invalid 'line_key_prefix' (must be string)", status_code=400)
    if not isinstance(prompt_hint_template, str):
        return make_api_response(error="Invalid 'prompt_hint_template' (must be string)", status_code=400)
    if target_placeholder not in prompt_hint_template:
         logging.warning(f"Target placeholder '{target_placeholder}' not found in prompt hint template: '{prompt_hint_template}'")
         # Proceed anyway, but the hint won't be dynamic

    db: Session = None
    new_lines_added = []
    try:
        db = next(get_db())

        # 1. Verify Script and Template Category exist
        script = db.query(models.VoScript).get(script_id)
        if not script:
            return make_api_response(error=f"Script {script_id} not found", status_code=404)
        
        category = db.query(models.VoScriptTemplateCategory).filter(
            # Ensure category belongs to the script's template
            models.VoScriptTemplateCategory.template_id == script.template_id, 
            models.VoScriptTemplateCategory.id == template_category_id,
            models.VoScriptTemplateCategory.is_deleted == False
        ).first()
        if not category:
            return make_api_response(error=f"Template Category {template_category_id} not found or does not belong to script's template", status_code=404)

        # 2. Determine starting order index for new lines within the category
        max_order_result = db.query(sa.func.max(models.VoScriptLine.order_index)).filter(
            models.VoScriptLine.vo_script_id == script_id,
            models.VoScriptLine.category_id == category.id
        ).scalar()
        start_order_index = (max_order_result or -1) + 1
        
        logging.info(f"Instantiating lines for category '{category.name}' (ID: {category.id}) in script {script_id}, starting order index: {start_order_index}")

        # 3. Loop through targets and create lines
        for i, target_name in enumerate(target_names):
            if not target_name or not target_name.strip():
                logging.warning(f"Skipping empty target name at index {i}.")
                continue
                
            target_name_clean = target_name.strip()
            
            # Sanitize target name for key (simple version)
            sanitized_target = "".join(c if c.isalnum() else '_' for c in target_name_clean).upper()
            
            # Generate line key and prompt hint
            final_line_key = f"{line_key_prefix}{sanitized_target}"
            final_prompt_hint = prompt_hint_template.replace(target_placeholder, target_name_clean)
            current_order_index = start_order_index + i

            # Check for duplicate key *within this script* before adding
            existing_line = db.query(models.VoScriptLine.id).filter(
                models.VoScriptLine.vo_script_id == script_id,
                models.VoScriptLine.line_key == final_line_key
            ).first()
            
            if existing_line:
                logging.warning(f"Skipping target '{target_name_clean}' - Line key '{final_line_key}' already exists in script {script_id}.")
                continue # Skip this target

            # Create the new line object
            new_line = models.VoScriptLine(
                vo_script_id=script_id,
                template_line_id=None, # Not linked to a specific template line
                category_id=category.id, # Link to the target category
                generated_text=None, # Starts empty
                status='pending', # Ready for generation
                is_locked=False, # Not locked by default
                line_key=final_line_key, 
                order_index=current_order_index,
                prompt_hint=final_prompt_hint
            )
            # Make sure status is set to pending (for debugging)
            logging.info(f"Creating line with key '{final_line_key}' and status='pending'")
            db.add(new_line)
            new_lines_added.append(new_line) # Keep track for response

        # 4. Commit all new lines if any were generated
        if not new_lines_added:
             return make_api_response(data={"message": "No new lines were added. Targets might be empty or keys already existed.", "lines_added_count": 0}, status_code=200)
        
        db.commit()
        logging.info(f"Successfully added {len(new_lines_added)} new lines for script {script_id} to category '{category.name}'.")

        # Optional: Return details of added lines?
        # For now, just return count
        return make_api_response(data={"message": f"Successfully added {len(new_lines_added)} lines.", "lines_added_count": len(new_lines_added)}, status_code=201)

    except Exception as e:
        if db and db.is_active: 
            try: db.rollback() 
            except: pass
        logging.exception(f"Error instantiating target lines for script {script_id}: {e}")
        return make_api_response(error="Failed to instantiate target lines.", status_code=500)
    finally:
        if db and db.is_active:
            db.close()

# --- Pydantic model for Chat Request Body ---
class ChatRequestBody(BaseModel):
    user_message: str
    current_context: Optional[Dict[str, Any]] = None # Keep this for focus
    image_base64_data: Optional[str] = None # NEW: Added for image uploads

# --- Pydantic model for Chat History Response Item ---
class ChatHistoryItem(BaseModel):
    role: str
    content: str
    timestamp: datetime

    class Config:
        from_attributes = True # Use from_attributes instead of orm_mode for Pydantic v2

# --- Chat Endpoints --- #

@vo_script_bp.route("/vo-scripts/<int:script_id>/chat", methods=["POST"])
def handle_chat_interaction(script_id: int):
    db = next(models.get_db())
    try:
        vo_script = db.query(models.VoScript).filter(models.VoScript.id == script_id).first()
        if not vo_script:
            return make_api_response(error="VO Script not found", status_code=404)

        if not request.is_json:
            return make_api_response(error="Request must be JSON", status_code=400)
        
        try:
            json_data = request.get_json()
            if json_data is None:
                return make_api_response(error="Invalid JSON body or missing Content-Type header", status_code=400)
            request_data = ChatRequestBody(**json_data)
        except ValidationError as e:
            return make_api_response(error=f"Invalid request body: {e.errors()}", status_code=400)
        except Exception as e_parse:
            return make_api_response(error="Invalid JSON body format", status_code=400)

        task = run_script_collaborator_chat_task.delay(
            script_id=script_id,
            user_message=request_data.user_message,
            initial_prompt_context_from_prior_sessions=None, # Pass None or empty list, it's ignored by task now
            current_context=request_data.current_context or {},
            image_base64_data=request_data.image_base64_data # NEW: Pass image data
        )
        
        current_app.logger.info(f"Dispatched AI Chat Collaborator task {task.id} for script ID {script_id}")
        return make_api_response(data={"task_id": task.id}, status_code=202)

    except Exception as e:
        current_app.logger.error(f"Error in chat interaction endpoint for script {script_id}: {e}", exc_info=True)
        return make_api_response(error="Internal server error during chat interaction.", status_code=500)
    finally:
        if db:
            db.close()

@vo_script_bp.route("/vo-scripts/<int:script_id>/chat/history", methods=["GET"])
def get_chat_history(script_id: int):
    """Endpoint to retrieve chat history for a specific VO Script."""
    db = next(models.get_db())
    try:
        history_records = db.query(models.ChatMessageHistory).filter(
            models.ChatMessageHistory.vo_script_id == script_id
        ).order_by(
            models.ChatMessageHistory.timestamp.asc() # Fetch in chronological order
        ).all()

        # Use Pydantic model for response structure
        # Ensure correct parsing from ORM object
        history_response = [ChatHistoryItem.model_validate(record).model_dump() for record in history_records]

        current_app.logger.info(f"Retrieved {len(history_response)} chat history messages for script {script_id}")
        return make_api_response(data=history_response, status_code=200)

    except Exception as e:
        current_app.logger.error(f"Error retrieving chat history for script {script_id}: {e}", exc_info=True)
        return make_api_response(error="Internal server error retrieving chat history.", status_code=500)
    finally:
        if db:
            db.close()

@vo_script_bp.route("/vo-scripts/<int:script_id>/chat/history", methods=["DELETE"])
def delete_chat_history(script_id: int):
    """Deletes all chat history for a specific VO Script."""
    db = next(models.get_db())
    try:
        # Verify the script exists (optional, but good practice)
        script = db.query(models.VoScript.id).filter(models.VoScript.id == script_id).first()
        if not script:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)

        # Delete chat messages associated with this script_id
        num_deleted = db.query(models.ChatMessageHistory).filter(
            models.ChatMessageHistory.vo_script_id == script_id
        ).delete(synchronize_session=False) # synchronize_session=False is common for bulk deletes
        
        db.commit()
        current_app.logger.info(f"Deleted {num_deleted} chat history messages for script {script_id}")
        return make_api_response(data={"message": f"Chat history for script {script_id} cleared successfully. {num_deleted} messages deleted."})

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Error deleting chat history for script {script_id}: {e}", exc_info=True)
        return make_api_response(error="Internal server error deleting chat history.", status_code=500)
    finally:
        if db:
            db.close()

# --- Pydantic model for Character Description Update --- #
class CharacterDescriptionUpdatePayload(BaseModel):
    new_description: str

# --- NEW Endpoint to commit staged character description update --- #
@vo_script_bp.route("/vo-scripts/<int:script_id>/character-description", methods=["PATCH"])
def commit_character_description_update(script_id: int):
    """Applies a new character description to the VO Script."""
    if not request.is_json:
        return make_api_response(error="Request must be JSON", status_code=400)
    
    try:
        payload = CharacterDescriptionUpdatePayload(**request.get_json())
    except ValidationError as e:
        return make_api_response(error=f"Invalid request body: {e.errors()}", status_code=400)
    except Exception as e_parse:
        current_app.logger.error(f"Error parsing description update request JSON: {e_parse}")
        return make_api_response(error="Invalid JSON body format", status_code=400)

    db: Session = None
    try:
        db = next(get_db())
        script = db.query(models.VoScript).get(script_id)
        if script is None:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)

        # Update the description
        script.character_description = payload.new_description
        db.commit()
        db.refresh(script)
        
        current_app.logger.info(f"Committed character description update for script {script_id}")
        # Return the updated script object (or relevant parts)
        # Use the same serialization as the GET endpoint for consistency
        script_data = model_to_dict(script) # Use existing serialization helper
        return make_api_response(data=script_data)

    except Exception as e:
        if db and db.is_active: db.rollback()
        current_app.logger.error(f"Error committing character description update for script {script_id}: {e}", exc_info=True)
        return make_api_response(error="Failed to update character description", status_code=500)
    finally:
        if db and db.is_active: db.close()

# --- Pydantic model for Scratchpad Note Response --- #
class ScriptNoteResponseItem(BaseModel):
    id: int
    vo_script_id: int
    category_id: Optional[int] = None
    line_id: Optional[int] = None
    title: Optional[str] = None
    text_content: str
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True # Pydantic v2

# --- NEW Endpoint to get scratchpad notes --- #
@vo_script_bp.route("/vo-scripts/<int:script_id>/scratchpad-notes", methods=["GET"])
def get_scratchpad_notes(script_id: int):
    """Fetches all scratchpad notes for a specific VO Script."""
    db = next(models.get_db())
    try:
        # Verify script exists (optional)
        script_exists = db.query(models.VoScript.id).filter(models.VoScript.id == script_id).scalar() is not None
        if not script_exists:
            return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)
        
        notes = db.query(models.ScriptNote).filter(
            models.ScriptNote.vo_script_id == script_id
        ).order_by(
            models.ScriptNote.updated_at.desc() # Show most recently updated first
        ).all()

        # Serialize using Pydantic model
        notes_response = [ScriptNoteResponseItem.model_validate(note).model_dump() for note in notes]

        current_app.logger.info(f"Retrieved {len(notes_response)} scratchpad notes for script {script_id}")
        return make_api_response(data=notes_response)

    except Exception as e:
        current_app.logger.error(f"Error retrieving scratchpad notes for script {script_id}: {e}", exc_info=True)
        return make_api_response(error="Internal server error retrieving scratchpad notes.", status_code=500)
    finally:
        if db:
            db.close()

# --- NEW Endpoint to delete a scratchpad note --- #
@vo_script_bp.route("/vo-scripts/<int:script_id>/scratchpad-notes/<int:note_id>", methods=["DELETE"])
def delete_scratchpad_note(script_id: int, note_id: int):
    """Deletes a specific scratchpad note."""
    db = next(models.get_db())
    try:
        # Find the note, ensuring it belongs to the correct script
        note = db.query(models.ScriptNote).filter(
            models.ScriptNote.id == note_id,
            models.ScriptNote.vo_script_id == script_id
        ).first()

        if not note:
            # Check if the script exists at all to give a better error
            script_exists = db.query(models.VoScript.id).filter(models.VoScript.id == script_id).scalar() is not None
            if not script_exists:
                return make_api_response(error=f"VO Script with ID {script_id} not found", status_code=404)
            else:
                 return make_api_response(error=f"Scratchpad note with ID {note_id} not found for script {script_id}", status_code=404)

        # Delete the note
        db.delete(note)
        db.commit()
        current_app.logger.info(f"Deleted scratchpad note {note_id} for script {script_id}")
        return make_api_response(data={"message": "Scratchpad note deleted successfully."}, status_code=200)

    except Exception as e:
        db.rollback()
        current_app.logger.error(f"Error deleting scratchpad note {note_id} for script {script_id}: {e}", exc_info=True)
        return make_api_response(error="Internal server error deleting scratchpad note.", status_code=500)
    finally:
        if db:
            db.close()